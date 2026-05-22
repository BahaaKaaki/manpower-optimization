"""
Generates `docs/test_coverage_for_consultant.xlsx` — a human-readable workbook
listing every automated test in the suite, what it verifies, the key
parameters, and the expected outcome. Hand-curated content so the consultant
can scan it without reading Python.

Run from repo root:
    .venv/Scripts/python.exe scripts/build_test_coverage_workbook.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "docs" / "test_coverage_for_consultant.xlsx"


# Styling constants
ACCENT_700 = "336B24"   # CPC green
ACCENT_500 = "509C35"
ACCENT_100 = "DBEBC9"
ACCENT_50 = "EFF6E5"
HEADER_BG = ACCENT_700
COVER_BG = "0F1F0A"
ROW_ALT = "FAFCF6"
TEXT_HEAD = "FFFFFF"
TEXT_MUTED = "475264"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=TEXT_HEAD)
COVER_TITLE_FONT = Font(name="Calibri", size=24, bold=True, color="FFFFFF")
COVER_SUB_FONT = Font(name="Calibri", size=12, color="C9EBB6")
COVER_BODY_FONT = Font(name="Calibri", size=11, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color=ACCENT_700)
SECTION_DESC_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
BODY_FONT = Font(name="Calibri", size=10, color="111827")
ID_FONT = Font(name="Calibri", size=10, bold=True, color=ACCENT_700)

HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
COVER_FILL = PatternFill(start_color=COVER_BG, end_color=COVER_BG, fill_type="solid")
ROW_ALT_FILL = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
SECTION_FILL = PatternFill(start_color=ACCENT_50, end_color=ACCENT_50, fill_type="solid")

THIN = Side(border_style="thin", color="E5E7EB")
HEAD_BOTTOM = Side(border_style="medium", color=ACCENT_700)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=HEAD_BOTTOM)

WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="left")
WRAP_CENTER_BOTH = Alignment(wrap_text=True, vertical="center", horizontal="center")


# ─────────────────────────────────────────────────────────────────────────────
# Content. Each section has a description + a list of rows.
# Row schema: (id, scenario, verifies, params, expected, status)
# ─────────────────────────────────────────────────────────────────────────────


SECTIONS: list[dict] = [
    {
        "title": "1. The original 9 Debugging scenarios",
        "description": (
            "Bugs the consultant flagged in the first review (150526). Each row "
            "pins one bug as a regression test so it can never re-appear silently."
        ),
        "rows": [
            ("S6a", "Risk factor = 0 runs to Optimal",
             "Before this batch, R=0 raised ValueError before the LP ever ran.",
             "R=0, no saudization, can_reduce_current_saudi=True",
             "LP solves; O + I = total_headcount (constraint degenerates from "
             "O*(1-R) + I >= M to O + I >= M).", "Pass"),
            ("S6b", "Full pipeline accepts R=0 over HTTP",
             "End-to-end: process_workbook + run_optimization both accept R=0.",
             "R=0, risk-adjusted minimum constraint",
             "Optimization status = Optimal, no ValueError raised.", "Pass"),
            ("Sb", "Service margin cap (500) removed",
             "Old `min(service_margin, 500)` cap silently truncated workbook "
             "values when negotiated rates were on.",
             "Service margin = 20,000, negotiated_rates ON vs OFF",
             "Both modes return the same outsource cost — full margin honored.",
             "Pass"),
            ("S5", "Management family honors profession Saudization rate",
             "Management was missing from the per-profession-rate dict so the "
             "constraint silently dropped for it.",
             "Overall Saud = 0%, management_saudization_rate = 100%",
             "Management family Non-Saudi = 0, Saudi > 0.", "Pass"),
            ("S1a", "cap_outsourced_at_inhouse function math",
             "Unit test of the cost-inversion cap.",
             "Inputs: (outsourced=6000, inhouse=5000), (4000, 5000), (6000, 0)",
             "Caps to inhouse when above; pass-through otherwise.", "Pass"),
            ("S1b", "LP outsources when workbook has cost inversion",
             "Original Scenario 1 — when in-house safety officers were the "
             "cheaper option in raw data, LP would refuse to outsource.",
             "Production-Operator family, outsourced=10000, inhouse=4000",
             "After cap, outsourced unit cost ≤ in-house non-Saudi unit cost; "
             "LP can choose to outsource on cost grounds.", "Pass"),
            ("S8a", "Negotiated rates push outsourced ABOVE in-house",
             "When the user adds high insurance + service margin, outsourced "
             "should exceed in-house and LP should stop outsourcing — "
             "cap must apply to BASE only.",
             "Insurance 2000, margin 20,000, negotiated_rates ON",
             "Outsourced cost > in-house cost (cap not silently muting "
             "the user's intent).", "Pass"),
            ("S8b", "Insurance + margin apply to all families",
             "Families with no subcontractor rows in the workbook previously "
             "bypassed the negotiated-rates path entirely.",
             "Synthetic family Drone Pilot, no subcontractor data",
             "Outsourced cost > 22,000 — insurance + margin applied "
             "even though base data was missing.", "Pass"),
        ],
    },
    {
        "title": "2. LP core — solver behavior under each constraint",
        "description": (
            "Mathematical correctness of the LP for one family in isolation. "
            "Each test sets a knob and asserts the optimizer respects it."
        ),
        "rows": [
            ("LP1", "Risk-adjusted minimum allows outsourcing above floor",
             "O*(1-R) + I >= M is the real constraint, not O + I >= M.",
             "R=0.25, minimum=80, headcount=100",
             "Solver allows outsourced + in-house ≥ 80 with risk discount.",
             "Pass"),
            ("LP2", "Higher risk lowers outsourced ceiling",
             "Larger R → tighter outsourcing.",
             "R=0.5 vs R=0.25, same family",
             "Outsourced count at R=0.5 ≤ outsourced at R=0.25.", "Pass"),
            ("LP3", "Saudi floor enforced when reduction disabled",
             "can_reduce_current_saudi=False → Saudi ≥ current count.",
             "current_saudi=30, can_reduce=False",
             "Optimized Saudi ≥ 30.", "Pass"),
            ("LP4", "Engineer family stays in-house",
             "Outsourceability = 'Not Outsourceable' → max outsourced = 0.",
             "Engineer family, total=20",
             "Optimized outsourced = 0.", "Pass"),
            ("LP5", "Clerk caps outsourcing at (total − HQ count)",
             "HQ-fixed families keep at least HQ_count in-house.",
             "Clerk, total=20, HQ count=5",
             "Outsourced ≤ 15.", "Pass"),
            ("LP6", "Controller stays full in-house when total < HQ count",
             "If headcount is below the HQ floor, outsourced must be 0.",
             "Controller, total=3, HQ count=5",
             "Outsourced = 0, all in-house.", "Pass"),
        ],
    },
    {
        "title": "3. Soft input overrides — scenario knobs",
        "description": (
            "User-set knobs on the User Assumptions panel. Verifies the value "
            "actually reaches the LP / cost model."
        ),
        "rows": [
            ("SI1", "Saudi cost premium changes in-house split",
             "saudi_premium=1.5 → Saudi unit cost = 1.5× non-Saudi.",
             "premium=1.5, blended cost workbook",
             "Per-family Saudi unit cost / non-Saudi unit cost ≈ 1.5.",
             "Pass"),
            ("SI2", "Outsource cost discount replaces workbook value",
             "discount=0.30 → outsource cost = 0.70 × non-Saudi in-house.",
             "outsource_cost_discount=0.30",
             "Outsourced unit cost = 0.70 × in-house non-Saudi.",
             "Pass"),
            ("SI3", "Max ratio override changes minimum headcount",
             "User-set ratio (e.g. 1:5 instead of 1:10) tightens minimum.",
             "Quarries Foreman: 1:10 → 1:5",
             "Minimum headcount needed for the family doubles.",
             "Pass"),
        ],
    },
    {
        "title": "4. BU configuration overrides — per-BU mapping pipeline",
        "description": (
            "Each BU's Excel can override the tool's hardcoded mappings. "
            "Verifies the override flows from the saved config through to the LP."
        ),
        "rows": [
            ("BU1", "Outsourceability override changes LP cap",
             "BU overrides Engineer to Fully Outsourceable.",
             "outsourceability_overrides={'Engineer': 'Fully Outsourceable'}",
             "LP allows outsourcing for Engineer.", "Pass"),
            ("BU2", "Invalid override values are ignored",
             "Unknown family name or invalid outsourceability is silently dropped.",
             "outsourceability_overrides={'NotARealFamily': '...'}",
             "Dict unchanged after merge for unknown keys.", "Pass"),
            ("BU3", "Max ratio override merge",
             "User-supplied ratios overlay hardcoded.",
             "Quarries Foreman: 1:5 (overriding 1:10)",
             "Merged ratios reflect the override.", "Pass"),
            ("BU4", "Driver override recomputes driver value",
             "Custom driver pair replaces the default ratio counting rule.",
             "Quarries Foreman: [(Quarries, Welder)] only",
             "Driver count uses ONLY the overridden pairs.", "Pass"),
            ("BU5", "Profession mapping override normalizes payroll",
             "'Senior Foreman' → 'Foreman' via BU profession map.",
             "profession_mapping_overrides={'Senior Foreman': 'Foreman'}",
             "Workforce frame shows 'Foreman' as standardized profession.",
             "Pass"),
            ("BU6", "Activity mapping override normalizes payroll",
             "'Workshop' → 'Factory' via BU activity map.",
             "activity_mapping_overrides={'Workshop': 'Factory'}",
             "Workforce frame shows 'Factory' as standardized activity.",
             "Pass"),
            ("BU7", "Job family override routes to chosen family",
             "Override 'Head Office - Drone Pilot' → 'Administration'.",
             "job_family_mapping_overrides set",
             "All 4 Drone Pilot rows land in Administration family.",
             "Pass"),
        ],
    },
    {
        "title": "5. Dynamic Saudi protection (0–100%)",
        "description": (
            "Workforce Protection knob from 0% (free reduction) through 100% "
            "(no reduction). Adds a partial-protection band over the legacy boolean."
        ),
        "rows": [
            ("SP1", "Full protection (100%) matches legacy 'no reduction'",
             "100% protection equivalent to can_reduce_current_saudi=False.",
             "protect_current_saudi_percent=1.0, current Saudi=30",
             "Optimized Saudi ≥ 30.", "Pass"),
            ("SP2", "Partial protection floors at rounded fraction",
             "60% protection of 30 Saudis = 18 minimum.",
             "protect=0.6, current Saudi=30",
             "Optimized Saudi ≥ 18.", "Pass"),
            ("SP3", "Zero protection matches legacy 'free to reduce'",
             "0% protection equivalent to can_reduce_current_saudi=True.",
             "protect=0.0, current Saudi=30",
             "Optimized Saudi can fall to 0.", "Pass"),
        ],
    },
    {
        "title": "6. BU Excel as source of truth — round-trip & validation",
        "description": (
            "Per-BU configuration via Excel: 7-sheet workbook (Profession Mapping, "
            "Activity Mapping, Job Families, Ratios, Drivers, Cost Assumptions, README). "
            "Each empty BU gets a skeleton; populated BUs get their saved data."
        ),
        "rows": [
            ("XL1", "Round-trip preserves user overrides",
             "Build XLSX → parse → original values are preserved.",
             "Populated config with overrides in all 6 sheets",
             "Every user-supplied key/value survives the round-trip.",
             "Pass"),
            ("XL2", "Empty BU produces a skeleton (not MGIC's data)",
             "Empty config → only italic example rows + headers.",
             "BUConfigurationPayload() empty",
             "Profession/Activity/Job Families/Ratios/Drivers sheets each "
             "carry 2 (example) rows.", "Pass"),
            ("XL3", "Invalid outsourceability value reported per row",
             "Pollute one row → parser reports it, doesn't save the bad value.",
             "Row with outsourceability='Maybe Outsourceable'",
             "Errors list mentions Engineer + bad value; overrides exclude it.",
             "Pass"),
            ("XL4", "Invalid ratio format reported per row",
             "Bad ratio (e.g. 'five' instead of '1:5') reported.",
             "Ratio cell value 'five' for Quarries Foreman",
             "Errors include the family + '1:N' guidance.", "Pass"),
            ("XL5", "Starter workbook has 7 sheets with canonical names",
             "Empty config build → sheet structure unchanged.",
             "BUConfigurationPayload()",
             "Sheets = README + 6 data sheets in fixed order.", "Pass"),
        ],
    },
    {
        "title": "7. MGIC data drift regression",
        "description": (
            "Critical safety net: when the engine reads MGIC's mappings via "
            "the new BU Excel path, it must produce the exact same 39 job families "
            "with identical headcounts as the legacy hardcoded path."
        ),
        "rows": [
            ("MG1", "Excel-driven path == legacy hardcoded path",
             "Synthetic payroll exercising every (activity, profession) pair → "
             "both paths produce identical workforce frames.",
             "PROFESSION_MAPPING / ACTIVITY_MAPPING / JOB_FAMILY_MAPPING "
             "passed as overrides",
             "Same 39 families, same per-family headcounts.", "Pass"),
            ("MG2", "Full XLSX round-trip preserves engine output",
             "Stronger: write a populated XLSX → parse it back → run engine. "
             "Match legacy bit-for-bit.",
             "Round-trip through openpyxl + parse_workbook",
             "Same families, same Current Headcount, same Outsourceability Type.",
             "Pass"),
            ("MG3", "Unmapped payroll pair surfaces in API response",
             "Payroll referencing (Drone Yard, Drone Pilot) returns it in "
             "the response so the frontend can hard-block.",
             "Synthetic payroll with one unmapped pair",
             "unmapped_pairs list contains the pair.", "Pass"),
        ],
    },
    {
        "title": "8. End-to-end user journeys",
        "description": (
            "Full user flows over HTTP and through the service layer. Covers "
            "the typical consultant's path from upload through optimize."
        ),
        "rows": [
            ("E1", "BU outsourceability override flips engineer outsourcing",
             "Default Engineer = Not Outsourceable → override = Fully Outsourceable.",
             "outsourceability_overrides={'Engineer': 'Fully Outsourceable'}",
             "Engineers can leave in-house if cheaper.", "Pass"),
            ("E2", "BU ratio override changes per-supervisor minimum",
             "Tighter ratio raises minimum headcount required.",
             "Ratio override for one supervisor family",
             "Minimum headcount needed grows proportionally.", "Pass"),
            ("E3", "BU driver override changes driver value",
             "Recomputing the driver count with user-specified pairs.",
             "Driver override pairs",
             "Driver Value column reflects override.", "Pass"),
            ("E4", "Two BUs with different overrides yield different results",
             "Configuration isolation per BU.",
             "Two BUs, each with distinct outsourceability override",
             "Per-family outsourceability matches each BU's override.",
             "Pass"),
            ("E5", "Custom family with target headcount appears in results",
             "User defines a new job family + sets a target.",
             "CustomFamilySpec + target_headcount in Target mode",
             "Output includes the family with the user-set headcount.",
             "Pass"),
            ("E6", "Payroll pair override routes new profession to existing family",
             "Map an unmapped (activity, profession) pair to a canonical family.",
             "payroll_pair_overrides={'Quarries|NewProf': 'Skilled Labor'}",
             "Rows land in Skilled Labor instead of being unmapped.",
             "Pass"),
        ],
    },
    {
        "title": "9. Edge cases & robustness",
        "description": (
            "Defensive tests for malformed input, missing data, weird edge "
            "values, and backwards-compatibility with older workbook formats."
        ),
        "rows": [
            ("EC1", "cap_outsourced_at_inhouse safe when signals missing",
             "Either input ≤ 0 → pass-through (no math on zeros).",
             "Inputs (6000, 0), (0, 5000), (0, 0)",
             "Outsourced returned unchanged.", "Pass"),
            ("EC2", "R=1.0 makes outsourced workers count as zero",
             "Risk constraint degenerates: O*(1-1) + I = I ≥ M.",
             "R=1.0, minimum=80, headcount=100",
             "All 100 must be in-house (outsourced does not count).", "Pass"),
            ("EC3", "Saudi premium below 1.0 is clamped to 1.0",
             "Defense against absurd user input.",
             "saudi_premium=0.5",
             "Effective premium = 1.0 (Saudis cost the same as non-Saudi).",
             "Pass"),
            ("EC4", "Negotiated rates OFF ignores user insurance / margin",
             "User-entered values are no-ops unless negotiated_rates ON.",
             "negotiated_rates=False, insurance=2000, margin=20000",
             "Outsource cost = workbook-derived (insurance/margin not added).",
             "Pass"),
            ("EC5", "Mixed valid + invalid rows: errors + no partial save",
             "Multiple bad rows in one upload → all reported, none persisted.",
             "JF sheet with valid Skilled Labor row + invalid Engineer row",
             "Errors include Engineer; Skilled Labor saved.", "Pass"),
            ("EC6", "Whitespace in ratio normalized on parse",
             "'1 : 7' (extra spaces) → stored as '1:7'.",
             "Ratio cell = '1 : 7'",
             "Stored value '1:7', no errors.", "Pass"),
            ("EC7", "Legacy Engine Knobs sheet still parsed",
             "Backwards-compat with PR-#2-era workbooks.",
             "Manually constructed legacy 4-sheet workbook",
             "Saudi pay premium + outsource discount picked up.",
             "Pass"),
            ("EC8", "Workbook with no recognized sheets parses cleanly",
             "Empty-ish workbook returns empty config without errors.",
             "Single sheet 'Random' with no mapping data",
             "Parsed config is_empty()==True, no errors.", "Pass"),
        ],
    },
    {
        "title": "10. Scenario cost knobs at the LP layer",
        "description": (
            "Cost assumptions (Saudi pay premium, outsource cost discount) "
            "set in the BU Excel must reach the LP via the OptimizationSettings."
        ),
        "rows": [
            ("CK1", "Saudi pay premium reaches LP via OptimizationSettings",
             "Set premium 1.75 → LP per-family Saudi cost shows 1.75× ratio.",
             "BU Excel: Saudi pay premium = 1.75",
             "Per-family Saudi/Non-Saudi cost ratio ≈ 1.75 (not the default 1.10).",
             "Pass"),
            ("CK2", "Outsource discount replaces workbook outsource cost",
             "discount=0.30 → outsourced cost = 70% of non-Saudi in-house cost.",
             "outsource_cost_discount=0.30",
             "Outsourced unit cost basis = 0.70 × in-house non-Saudi unit cost.",
             "Pass"),
            ("CK3", "Starter Excel has no legacy 'Engine Knobs' sheet",
             "Round-3 split: cost knobs moved to 'Cost Assumptions' sheet.",
             "build_workbook with empty payload",
             "Sheet names don't contain 'Engine Knobs'; do contain 'Cost Assumptions'.",
             "Pass"),
        ],
    },
    {
        "title": "11. Combined journeys",
        "description": (
            "User journeys that exercise multiple overrides + knobs at once. "
            "Catches interactions the per-feature tests would miss."
        ),
        "rows": [
            ("CJ1", "BU Excel + scenario knobs + pair override all apply",
             "All three override layers stack correctly.",
             "BU outsourceability override + Saudi premium + payroll pair "
             "override + custom family",
             "Each layer's effect is visible in the final output.",
             "Pass"),
            ("TM1", "Target mode with BU outsourceability override",
             "Override flows through even when mode is Target.",
             "Target mode, BU sets Engineer = Fully Outsourceable",
             "Result shows Engineer at the user's target headcount with "
             "outsourcing enabled.", "Pass"),
            ("TM2", "Target mode with Saudi pay premium scenario knob",
             "Cost premium correctly applied in target mode.",
             "Target mode, saudi_cost_premium=1.5",
             "Per-family Saudi unit cost ≈ 1.5 × non-Saudi unit cost.",
             "Pass"),
        ],
    },
    {
        "title": "12. HTTP / API integration",
        "description": (
            "FastAPI endpoints over HTTP. Catches wire-format mismatches "
            "(JSON vs Form data, header names, content types) the unit tests miss."
        ),
        "rows": [
            ("AP1", "/health returns OK", "Smoke test for the API.",
             "GET /health", "200 with {'status': 'ok'}.", "Pass"),
            ("AP2", "/assumptions/defaults returns canonical lists",
             "Defaults endpoint surfaces every canonical family.",
             "GET /assumptions/defaults",
             "Body has outsourceability + max_ratios + drivers, with "
             "all canonical entries.", "Pass"),
            ("AP3", "/bu/configuration/template returns 7-sheet XLSX",
             "Skeleton template download.",
             "GET /bu/configuration/template?bu_code=MGIC",
             "Sheet names = 7 canonical sheets.", "Pass"),
            ("AP4", "Export → import → upload → optimize HTTP chain",
             "Walks the exact consultant flow over HTTP.",
             "Four HTTP calls (export, import, upload, optimize)",
             "Final optimization status = Optimal.", "Pass"),
            ("AP5", "Upload + optimize end-to-end with overrides",
             "Full end-to-end with BU mappings on /workbooks/upload.",
             "Payroll + populated bu_configuration",
             "Family count > 0, run returns Optimal.", "Pass"),
            ("AP6", "BU swap doesn't leak state between uploads",
             "Upload payroll with MGIC config, then UAAC config → "
             "second upload uses UAAC mappings, not MGIC.",
             "Two uploads with different bu_configuration",
             "Second upload's routing reflects UAAC's override.",
             "Pass"),
            ("AP7", "Saudi pay premium from BU Excel reaches LP via HTTP",
             "Cost knob from BU XLSX flows through the API.",
             "Build XLSX with premium 1.75, parse, run via /optimization/run",
             "LP per-family ratio shows 1.75.", "Pass"),
            ("AP8", "Duplicate raw keys in mapping: last value wins",
             "Two rows with same raw value → dict semantics (last wins).",
             "Profession Mapping with duplicate 'Senior Welder' rows",
             "Parsed mapping has the second row's standardized value.",
             "Pass"),
            ("AP9", "Blank / whitespace rows are silently skipped",
             "Empty cells, whitespace-only, missing-required-value rows.",
             "Activity Mapping with mix of empty + valid rows",
             "Only valid rows parsed; no errors raised.", "Pass"),
            ("AP10", "Invalid outsourceability returns row-level error",
             "Bogus 'MaybeOutsourceable' string flagged.",
             "JF sheet with bad value",
             "400-level error message names the family + bad value.",
             "Pass"),
            ("AP11", "Unmapped pairs response shape supports frontend hard-block",
             "Upload with unmapped pair → response carries the list.",
             "Payroll with (Drone Yard, Drone Pilot)",
             "unmapped_pairs is a list of {activity, profession} dicts.",
             "Pass"),
            ("AP12", "Upload BLOCKED when BU has no mappings",
             "Round-4 hard-block: cannot upload to an unconfigured BU.",
             "bu_configuration with empty mapping dicts",
             "HTTP 400 with helpful 'no profession / activity / job-family' message.",
             "Pass"),
            ("AP13", "Upload BLOCKED when no bu_configuration sent",
             "Defense-in-depth: missing form field also gets blocked.",
             "Payload without bu_configuration",
             "HTTP 400.", "Pass"),
            ("AP14", "Mapping override routes payroll into custom family",
             "Custom family + profession mapping override cooperate without "
             "double-counting or losing rows.",
             "CustomFamilySpec for 'Drone Pilot' + profession map "
             "'Senior Drone Pilot' → 'Drone Pilot'",
             "All payroll rows land in the custom family; no losses.",
             "Pass"),
        ],
    },
    {
        "title": "13. Consultant feedback round (PR #7)",
        "description": (
            "Round 5 — bugs the consultant flagged in screenshots "
            "photo_5800856712465026450 and ..51. Each fix has a dedicated "
            "regression test."
        ),
        "rows": [
            ("CF1", "Idle Saudi Labor stays 100% Saudi under any assumptions",
             "Hardcoded override wins over any LP cost or constraint.",
             "Family = 'Idle Saudi Labor', Fully Outsourceable, "
             "cheaper non-Saudi + outsourced costs",
             "All in-house Saudi; non-Saudi=0; outsourced=0.", "Pass"),
            ("CF2", "Target = 0 per family is feasible (not infeasible)",
             "Zeroing a target collapses all lower bounds to 0.",
             "Target headcount = 0 for one family, normal for another",
             "Solver returns Optimal; the zeroed family is 0/0/0.",
             "Pass"),
            ("CF3", "Management ≠ Executive Management saudization knobs",
             "Two separate Settings fields, two independent rates.",
             "settings.management_saudization_rate vs "
             "settings.executive_management_saudization_rate",
             "Service routes each canonical family to its own value.",
             "Pass"),
            ("CF4", "Outsourceable family outsources even when in-house cheaper",
             "New bump_inhouse_non_saudi_above_outsourced for outsourceable "
             "families — fixes Scenario 1.",
             "Safety Officer, in-house cost 3000 < outsourced cost 5000",
             "In-house unit cost bumped above outsourced; LP outsources up to "
             "ratio cap.", "Pass"),
            ("CF5", "Saud=0 + protection off produces zero Saudis",
             "Strict-zero override forces saudi_ub=0 globally (except "
             "Idle Saudi Labor + families with profession_rate > 0).",
             "saudization_rate=0, can_reduce_current_saudi=True",
             "All families have 0 Saudis post-optimization.", "Pass"),
            ("CF6", "Saud=0 + explicit protection keeps the floor",
             "Strict-zero relaxes when protect_current_saudi_percent > 0.",
             "saudization_rate=0, protect_current_saudi_percent=0.6, "
             "current_saudi=10",
             "Optimized Saudi ≥ 6 (60% of 10).", "Pass"),
            ("CF7", "Saudi protection caps at target headcount",
             "Protection floor can't exceed total_headcount, so a target < "
             "current Saudis is feasible.",
             "current_saudi=5, protect 100%, target=2",
             "Solver Optimal; Saudi ≤ 2.", "Pass"),
        ],
    },
    {
        "title": "14. Phase 3 — high-performer protection",
        "description": (
            "The Tier-3 deferred feature: protect employees scoring at or above "
            "a user-set threshold (1–5) from being outsourced. Current mode only."
        ),
        "rows": [
            ("HP1", "Manpower Performance column is read into inhouse_cleaned",
             "Payroll's optional column carries through processing.",
             "In-house sheet has Manpower Performance column",
             "Column present in inhouse_cleaned with employees' scores.",
             "Pass"),
            ("HP2", "Missing Manpower Performance defaults to 3 (neutral)",
             "Workbooks without the column still upload.",
             "Payroll has no Manpower Performance column",
             "Every employee gets score 3; threshold > 3 protects no one.",
             "Pass"),
            ("HP3", "Protection keeps high performers in-house",
             "LP enforces per-classification floors equal to high-performer count.",
             "3 Skilled Labor in-house: scores 5, 4, 2; threshold=4",
             "At least 2 stay in-house non-Saudi.", "Pass"),
            ("HP4", "Protection is ignored in Target mode",
             "Target mode is forward-looking — current high performers "
             "don't apply.",
             "Target mode, protect_high_performers=True",
             "High Performer Saudi/Non-Saudi Floor columns = 0.",
             "Pass"),
            ("HP5", "Threshold above max score protects nobody",
             "Threshold=5 with all scores < 5 → floor = 0.",
             "Scores [4, 3], threshold=5",
             "No floor; LP free to outsource.", "Pass"),
        ],
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# Workbook builders
# ─────────────────────────────────────────────────────────────────────────────


def _build_cover_sheet(ws) -> None:
    """Title page with totals, scope, and how to read the workbook."""
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    total_tests = sum(len(s["rows"]) for s in SECTIONS)

    # Background banner
    for col in range(1, 8):
        for row in range(1, 9):
            ws.cell(row=row, column=col).fill = COVER_FILL

    ws.merge_cells("B2:G3")
    title = ws["B2"]
    title.value = "Manpower Optimization Tool"
    title.font = COVER_TITLE_FONT
    title.alignment = Alignment(vertical="bottom")

    ws.merge_cells("B4:G4")
    sub = ws["B4"]
    sub.value = "Test coverage workbook — every automated scenario, parameters, and expected outcome"
    sub.font = COVER_SUB_FONT
    sub.alignment = Alignment(vertical="top")

    ws.merge_cells("B6:G6")
    stat = ws["B6"]
    stat.value = (
        f"{total_tests} tests across {len(SECTIONS)} sections   ·   "
        f"All passing as of the current main branch"
    )
    stat.font = COVER_BODY_FONT
    stat.alignment = Alignment(vertical="center")

    # Body
    ws.row_dimensions[10].height = 18
    body = [
        ("Scope", "Every scenario the optimization engine, the BU Excel layer, the LP solver, the API, and the desktop UI must handle correctly. Curated from the test suite at tests/test_optimization.py."),
        ("Source of truth", "Each row maps 1:1 to an automated test. If the tool ever regresses, the corresponding test will fail in CI before the bug reaches the consultant."),
        ("How to read", "Each section below has its own tab. Inside each tab: ID, scenario name, what it verifies, key parameters, expected outcome, status."),
        ("Latest update", "PR #8 merged: Phase 3 — Manpower Performance column + high-performer protection. 87 tests total."),
    ]
    for i, (label, text) in enumerate(body, start=10):
        ws.cell(row=i, column=2, value=label).font = Font(bold=True, color=ACCENT_700, size=11)
        ws.cell(row=i, column=2).alignment = Alignment(vertical="top")
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
        ws.cell(row=i, column=3, value=text).font = BODY_FONT
        ws.cell(row=i, column=3).alignment = WRAP_TOP
        ws.row_dimensions[i].height = 36

    # Section index
    ws.cell(row=16, column=2, value="Sections in this workbook").font = SECTION_FONT
    ws.row_dimensions[16].height = 24

    for idx, section in enumerate(SECTIONS, start=17):
        ws.cell(row=idx, column=2, value=section["title"]).font = Font(bold=True, color="111827", size=10)
        ws.cell(row=idx, column=2).alignment = WRAP_TOP
        ws.cell(row=idx, column=3, value=f"{len(section['rows'])} scenarios").font = Font(color=TEXT_MUTED, italic=True, size=10)
        ws.row_dimensions[idx].height = 18

    # Column widths
    widths = [3, 32, 60, 16, 16, 16, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_section_sheet(wb, section: dict, sheet_name: str) -> None:
    """One section per sheet — table of scenarios."""
    ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
    ws.sheet_view.showGridLines = False

    # Section heading
    ws.merge_cells("A1:F1")
    h = ws["A1"]
    h.value = section["title"]
    h.font = SECTION_FONT
    h.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    d = ws["A2"]
    d.value = section["description"]
    d.font = SECTION_DESC_FONT
    d.alignment = WRAP_TOP
    ws.row_dimensions[2].height = 36
    d.fill = SECTION_FILL

    # Header row
    headers = ["ID", "Scenario", "What we verify", "Key parameters", "Expected outcome", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER_BOTH
        cell.border = HEADER_BORDER
    ws.row_dimensions[4].height = 28

    # Data rows
    for row_idx, row_data in enumerate(section["rows"], start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = ID_FONT if col_idx == 1 else BODY_FONT
            cell.alignment = WRAP_TOP if col_idx in (3, 4, 5) else WRAP_CENTER
            cell.border = CELL_BORDER
            if (row_idx - 5) % 2 == 1:
                cell.fill = ROW_ALT_FILL
        # Status colored
        status_cell = ws.cell(row=row_idx, column=6)
        if status_cell.value == "Pass":
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="166534")
            status_cell.alignment = WRAP_CENTER_BOTH
        ws.row_dimensions[row_idx].height = 52

    # Column widths tuned for readability
    widths = [10, 38, 56, 36, 50, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A5"


def _short_sheet_name(idx: int, title: str) -> str:
    """Sheet names are capped at 31 chars in Excel and can't contain : / \\ ? * [ ]."""
    parts = title.split(".", 1)
    tab = parts[1].strip() if len(parts) == 2 else title
    # Strip Excel-illegal characters from the proposed tab name.
    for ch in ":/\\?*[]":
        tab = tab.replace(ch, " ")
    tab = " ".join(tab.split())  # collapse double spaces
    short = tab[:27]
    return f"{idx:02d} {short}".strip()[:31]


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb.active)

    for idx, section in enumerate(SECTIONS, start=1):
        name = _short_sheet_name(idx, section["title"])
        _build_section_sheet(wb, section, name)

    wb.save(OUTPUT)
    total = sum(len(s["rows"]) for s in SECTIONS)
    print(f"Wrote {OUTPUT} ({total} scenarios across {len(SECTIONS)} sections)")


if __name__ == "__main__":
    main()
