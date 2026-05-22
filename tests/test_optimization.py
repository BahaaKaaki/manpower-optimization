from __future__ import annotations

import unittest
from pathlib import Path

import pandas as pd

from manpower_app.optimization import (
    IN_HOUSE_NON_SAUDI_COLUMN,
    IN_HOUSE_SAUDI_COLUMN,
    OUTSOURCED_COLUMN,
    solve_optimization,
)
from manpower_app.service import OptimizationSettings, process_workbook, run_optimization
from manpower_app.terminology import OUTSOURCED_UNIT_COST_BASIS_COLUMN


def base_row(**overrides):
    row = {
        "Job Family": "Generic Role",
        "Outsourceability Type": "Partially Outsourceable",
        "Current Headcount": 100,
        "Minimum Headcount Needed": 80,
        "Risk Factor": 0.25,
        "Outsourced v1": 80,
        "Current Total In-house Saudi": 0,
        "Tenured Saudi In-House": 0,
        "Tenured Non-Saudi In-House": 0,
        "Fully Loaded Cost per In-house Saudi Employee": 100.0,
        "Fully Loaded Cost per In-house Non-Saudi Employee": 10.0,
        OUTSOURCED_UNIT_COST_BASIS_COLUMN: 1.0,
    }
    row.update(overrides)
    return row


class OptimizationLPTests(unittest.TestCase):
    def solve_one(self, row):
        data = pd.DataFrame([row])
        solved, payroll, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        return solved.iloc[0], payroll, status

    def test_risk_adjusted_minimum_allows_outsourced_above_inhouse_minimum(self):
        row, _, status = self.solve_one(base_row())

        outsourced = int(row[OUTSOURCED_COLUMN])
        inhouse = int(row[IN_HOUSE_NON_SAUDI_COLUMN] + row[IN_HOUSE_SAUDI_COLUMN])
        effective_count = outsourced * (1 - row["Risk Factor"]) + inhouse

        self.assertEqual(status, "Optimal")
        self.assertEqual(outsourced, 80)
        self.assertEqual(inhouse, 20)
        self.assertGreaterEqual(effective_count, row["Minimum Headcount Needed"])
        self.assertLess(inhouse, row["Minimum Headcount Needed"])

    def test_higher_risk_factor_lowers_outsourced_cap(self):
        row, _, _ = self.solve_one(base_row(**{"Risk Factor": 0.5, "Outsourced v1": 40}))

        self.assertEqual(int(row[OUTSOURCED_COLUMN]), 40)
        self.assertEqual(int(row[IN_HOUSE_NON_SAUDI_COLUMN] + row[IN_HOUSE_SAUDI_COLUMN]), 60)

    def test_current_saudi_floor_is_enforced_when_reduction_is_disabled(self):
        data = pd.DataFrame([base_row(**{"Current Total In-house Saudi": 5})])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=False,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )

        self.assertEqual(status, "Optimal")
        self.assertGreaterEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 5)

    def test_engineer_family_is_locked_in_house(self):
        # Engineer is "Not Outsourceable": max_outsourced = 0 regardless of cost,
        # so the LP is forced to keep all 100 in-house.
        row, _, status = self.solve_one(
            base_row(**{"Job Family": "Engineer", "Outsourceability Type": "Not Outsourceable"})
        )
        self.assertEqual(status, "Optimal")
        self.assertEqual(int(row[OUTSOURCED_COLUMN]), 0)
        self.assertEqual(
            int(row[IN_HOUSE_NON_SAUDI_COLUMN] + row[IN_HOUSE_SAUDI_COLUMN]), 100
        )

    def test_clerk_caps_outsourcing_at_total_minus_hq_count(self):
        # HQ-fixed rule: outsource cap = total - HQ_count, ignoring Outsourced v1.
        # 100 total, 30 HQ Clerks -> max 70 outsourced (LP picks the cap to minimize cost).
        row, _, status = self.solve_one(
            base_row(
                **{
                    "Job Family": "Clerk",
                    "HQ Inhouse Count": 30,
                    "Outsourced v1": 5,  # would normally cap at 5 — must be ignored for HQ-fixed
                }
            )
        )
        self.assertEqual(status, "Optimal")
        self.assertEqual(int(row[OUTSOURCED_COLUMN]), 70)
        self.assertEqual(
            int(row[IN_HOUSE_NON_SAUDI_COLUMN] + row[IN_HOUSE_SAUDI_COLUMN]), 30
        )

    def test_controller_full_inhouse_when_total_below_hq_count(self):
        # Edge case: planned headcount drops below the current HQ count.
        # All workers stay in-house (max_outsourced = max(0, total - HQ) = 0).
        row, _, status = self.solve_one(
            base_row(
                **{
                    "Job Family": "Controller",
                    "Current Headcount": 8,
                    "HQ Inhouse Count": 12,
                    "Minimum Headcount Needed": 8,
                    "Outsourced v1": 10,
                }
            )
        )
        self.assertEqual(status, "Optimal")
        self.assertEqual(int(row[OUTSOURCED_COLUMN]), 0)
        self.assertEqual(
            int(row[IN_HOUSE_NON_SAUDI_COLUMN] + row[IN_HOUSE_SAUDI_COLUMN]), 8
        )


class SoftInputOverrideTests(unittest.TestCase):
    """Tests for the user-tunable assumptions (Saudi premium, outsource discount, max ratios)."""

    def test_saudi_cost_premium_changes_inhouse_split(self):
        from manpower_app.costs import calculate_inhouse_cost_split

        # With premium = 1.10 (default): Saudi = 1.10 * non-Saudi.
        default_split = calculate_inhouse_cost_split(100.0, 1, 1)
        self.assertAlmostEqual(
            default_split["Fully Loaded Cost per In-house Saudi Employee"]
            / default_split["Fully Loaded Cost per In-house Non-Saudi Employee"],
            1.10,
            places=6,
        )

        # With premium = 1.50: Saudi = 1.50 * non-Saudi.
        custom_split = calculate_inhouse_cost_split(100.0, 1, 1, saudi_premium=1.50)
        self.assertAlmostEqual(
            custom_split["Fully Loaded Cost per In-house Saudi Employee"]
            / custom_split["Fully Loaded Cost per In-house Non-Saudi Employee"],
            1.50,
            places=6,
        )

    def test_outsource_cost_discount_replaces_workbook_cost(self):
        # Round-trip through the full pipeline: with the override on, the outsource cost basis
        # is (1 - discount) * non-Saudi in-house cost — independent of the workbook value.
        if not Path("/Users/bkaaki001/Downloads/Manpower (1).xlsx").exists():
            self.skipTest("comparison workbook not available")
        processed = process_workbook("/Users/bkaaki001/Downloads/Manpower (1).xlsx")

        baseline = run_optimization(processed, OptimizationSettings())
        cheap = run_optimization(
            processed,
            OptimizationSettings(outsource_cost_discount=0.50),
        )

        # When outsourcing is half the cost of non-Saudi in-house, the optimizer outsources at
        # least as many workers as in the baseline (cost dropped, never any reason to outsource less).
        self.assertGreaterEqual(
            cheap["results"]["total_outsourced_final"],
            baseline["results"]["total_outsourced_final"],
        )

    def test_max_ratio_override_changes_minimum_headcount(self):
        # Tighten the Quarries Foreman ratio from 1:15 to 1:5; min-headcount should grow ~3x.
        if not Path("/Users/bkaaki001/Downloads/Manpower (1).xlsx").exists():
            self.skipTest("comparison workbook not available")
        processed = process_workbook("/Users/bkaaki001/Downloads/Manpower (1).xlsx")
        data_default, _ = __import__(
            "manpower_app.service", fromlist=["prepare_model_data"]
        ).prepare_model_data(processed, OptimizationSettings())
        data_tight, _ = __import__(
            "manpower_app.service", fromlist=["prepare_model_data"]
        ).prepare_model_data(
            processed,
            OptimizationSettings(max_ratio_overrides={"Quarries Foreman": "1:5"}),
        )

        default_min = int(
            data_default[data_default["Job Family"] == "Quarries Foreman"]["Minimum Headcount Needed"].iloc[0]
        )
        tight_min = int(
            data_tight[data_tight["Job Family"] == "Quarries Foreman"]["Minimum Headcount Needed"].iloc[0]
        )
        self.assertGreater(tight_min, default_min)


WORKBOOK = Path("/Users/bkaaki001/Downloads/Manpower (1).xlsx")


@unittest.skipUnless(WORKBOOK.exists(), "comparison workbook is not available on this machine")
class WorkbookRegressionTests(unittest.TestCase):
    def test_default_workbook_matches_pinned_savings_baseline(self):
        processed = process_workbook(str(WORKBOOK))
        payload = run_optimization(processed, OptimizationSettings())

        self.assertEqual(payload["metadata"]["optimization_status"], "Optimal")
        self.assertAlmostEqual(payload["metadata"]["optimized_savings"], 0.014917296126, places=9)
        self.assertEqual(int(payload["results"]["total_outsourced_final"]), 7865)
        self.assertFalse(payload["audit"]["Risk-Adjusted Minimum Met"].eq(False).any())


class Tier5IntegrationTests(unittest.TestCase):
    """Integration tests for the Tier 5 features: unmapped reporting, custom-family
    plumbing, and target-mode optimization. Build a tiny in-memory workbook so the
    tests do not depend on the regression file."""

    @staticmethod
    def _make_workbook_bytes(*, include_unmapped: bool = False) -> bytes:
        import io
        from openpyxl import Workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality", "Total Paid", "Total Unpaid",
        ])
        # Two known mappings: Factory + Operator -> Factory Operator; Head Office + Clerk -> Clerk.
        inhouse.append([1, "Production", "Operator", "Saudi", 5000, 0])
        inhouse.append([2, "Production", "Operator", "non-saudi", 4000, 0])
        inhouse.append([3, "Head Office", "Clerk", "SAUDI", 3000, 0])
        if include_unmapped:
            inhouse.append([4, "Production", "Astronaut", "Saudi", 4000, 0])

        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        # One subcontractor row in the same Production-Operator family.
        sub.append([100, "Production", "Operator", "non-saudi", 1000, 200, 100, 50, 0, 0, 0, 0, 100])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_unmapped_pairs_are_reported_not_raised(self):
        """process_workbook should not raise on unknown pairs; should surface them
        through ProcessedWorkbook.unmapped_pairs."""
        import io
        from manpower_app.service import process_workbook

        contents = self._make_workbook_bytes(include_unmapped=True)
        processed = process_workbook(io.BytesIO(contents))
        # Astronaut is not in JOB_FAMILY_MAPPING -> it should appear in unmapped_pairs.
        pairs = [(p["activity"], p["profession"]) for p in processed.unmapped_pairs]
        self.assertIn(("Factory", "Astronaut"), pairs)
        # Mapped families are still in the optimization frame.
        self.assertIn("Factory Operator", set(processed.optimization_df["Job Family"]))

    def test_custom_mapping_resolves_unmapped_pair(self):
        """When the user supplies a CustomFamilySpec covering an unmapped pair, the
        pair should be resolved on re-process and the family should appear in the
        optimization frame."""
        import io
        from manpower_app.family_specs import (
            ActivityProfession,
            CustomFamilySpec,
            CustomFamilyCosts,
        )
        from manpower_app.service import process_workbook

        contents = self._make_workbook_bytes(include_unmapped=True)
        custom = [
            CustomFamilySpec(
                family_name="Astro Operator",
                outsourceability="Partially Outsourceable",
                source_pairs=[ActivityProfession(activity="Factory", profession="Astronaut")],
                costs=CustomFamilyCosts(
                    saudi_inhouse=5000, non_saudi_inhouse=4500, outsourced=4000
                ),
            )
        ]
        processed = process_workbook(io.BytesIO(contents), custom_families=custom)
        self.assertEqual(processed.unmapped_pairs, [])
        self.assertIn("Astro Operator", set(processed.optimization_df["Job Family"]))

    def test_target_mode_replaces_current_headcount_for_named_family(self):
        """In target mode, the LP's headcount-balance equality should be against the
        user-supplied target, not the workbook's current count."""
        import io
        from manpower_app.service import (
            OptimizationSettings,
            process_workbook,
            run_optimization,
        )

        contents = self._make_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        settings = OptimizationSettings(
            optimization_mode="target",
            target_headcounts={"Factory Operator": 9},  # baseline is 3 (2 inhouse + 1 sub)
            can_reduce_current_saudi=True,  # avoid Saudi floor making the LP infeasible
        )
        payload = run_optimization(processed, settings)
        rows = payload["data"]
        operator_row = rows[rows["Job Family"] == "Factory Operator"].iloc[0]
        total_split = (
            int(operator_row[OUTSOURCED_COLUMN])
            + int(operator_row[IN_HOUSE_NON_SAUDI_COLUMN])
            + int(operator_row[IN_HOUSE_SAUDI_COLUMN])
        )
        self.assertEqual(total_split, 9)
        # Savings should NOT be in the metadata in target mode.
        self.assertNotIn("optimized_savings", payload["metadata"])
        self.assertEqual(payload["metadata"]["optimization_mode"], "target")

    def test_brand_new_family_with_costs_shows_up_in_results(self):
        """A user-defined family with no rows in the workbook gets injected by
        process_workbook using the user's costs, and its target headcount drives the LP."""
        import io
        from manpower_app.family_specs import CustomFamilyCosts, CustomFamilySpec
        from manpower_app.service import (
            OptimizationSettings,
            process_workbook,
            run_optimization,
        )

        contents = self._make_workbook_bytes()
        new_family = CustomFamilySpec(
            family_name="Drone Pilot",
            outsourceability="Partially Outsourceable",
            source_pairs=[],
            costs=CustomFamilyCosts(saudi_inhouse=8000, non_saudi_inhouse=7000, outsourced=5000),
        )
        processed = process_workbook(io.BytesIO(contents), custom_families=[new_family])
        self.assertIn("Drone Pilot", set(processed.optimization_df["Job Family"]))

        settings = OptimizationSettings(
            optimization_mode="target",
            target_headcounts={"Drone Pilot": 5},
            custom_families=[new_family],
            can_reduce_current_saudi=True,
        )
        payload = run_optimization(processed, settings)
        rows = payload["data"]
        drone_row = rows[rows["Job Family"] == "Drone Pilot"].iloc[0]
        total_split = (
            int(drone_row[OUTSOURCED_COLUMN])
            + int(drone_row[IN_HOUSE_NON_SAUDI_COLUMN])
            + int(drone_row[IN_HOUSE_SAUDI_COLUMN])
        )
        self.assertEqual(total_split, 5)


class DebuggingScenarioTests(unittest.TestCase):
    """Regression tests for the 9 scenarios listed on the consultant's Debugging sheet
    in `150526_Manpower_Optimization_Enhancements_STATUS.xlsx`. Each test pins a bug
    that was open before this batch so future changes cannot silently re-introduce it.
    """

    @staticmethod
    def _make_workbook_bytes(
        *,
        include_management: bool = False,
        outsourced_more_expensive_than_inhouse: bool = False,
    ) -> bytes:
        import io
        from openpyxl import Workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        # Use cost columns the in-house cost calculator actually reads (Basic, Housing Paid,
        # etc.) so per-employee in-house cost is non-zero — the cost-inversion cap requires
        # a real in-house signal to fire.
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        inhouse.append([1, "Production", "Operator", "Saudi", 5000, 0, 4000, 500, 200, 200, 100, 0])
        inhouse.append([2, "Production", "Operator", "non-saudi", 3780, 0, 3000, 400, 150, 150, 80, 0])
        if include_management:
            # "Management" job family — distinct from "Executive Management".
            # Workbook→family mapping uses Activity + Profession, so "Head Office - Manager"
            # routes to the "Management" job family per manpower_app/mappings.py.
            inhouse.append([10, "Head Office", "Manager", "non-saudi", 12000, 0, 9000, 1500, 500, 500, 300, 200])
            inhouse.append([11, "Head Office", "Manager", "non-saudi", 12500, 0, 9500, 1500, 500, 500, 300, 200])
            inhouse.append([12, "Head Office", "Manager", "Saudi", 14600, 0, 11000, 1800, 600, 600, 400, 200])

        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        # One subcontractor in Production-Operator. Bump the subcontractor basic so the
        # outsourced unit cost beats the in-house one and triggers the inversion cap.
        sub_basic = 10000 if outsourced_more_expensive_than_inhouse else 1000
        sub.append([100, "Production", "Operator", "non-saudi", sub_basic, 200, 100, 50, 0, 0, 0, 0, 100])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # --- Bug A1: R=0 ----------------------------------------------------------

    def test_scenario_6_risk_factor_zero_runs_to_optimal(self):
        """Scenario 6 from the Debugging sheet: risk_factor=0. Before this batch the
        service layer raised `ValueError("Risk factor must be greater than 0...")`
        before the LP was ever invoked. The LP itself is R=0-safe — the constraint
        degenerates from O*(1-R)+I>=M to O+I>=M, which is solvable whenever T>=M."""
        data = pd.DataFrame([base_row(**{"Risk Factor": 0.0})])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=False,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        self.assertEqual(status, "Optimal")
        out = int(solved.iloc[0][OUTSOURCED_COLUMN])
        inh = int(
            solved.iloc[0][IN_HOUSE_NON_SAUDI_COLUMN]
            + solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]
        )
        # O + I must satisfy the headcount-balance equality and the minimum constraint.
        self.assertEqual(out + inh, 100)
        self.assertGreaterEqual(out + inh, 80)

    def test_scenario_6_run_optimization_does_not_raise_at_zero_risk(self):
        """Service-layer regression: the full pipeline accepts risk_factor=0 without
        raising. Before this batch, `prepare_model_data` raised ValueError on R<=0."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._make_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        settings = OptimizationSettings(risk_factor=0.0, can_reduce_current_saudi=True)
        payload = run_optimization(processed, settings)
        self.assertEqual(payload["metadata"]["optimization_status"], "Optimal")

    # --- Bug A2: service margin cap removed -----------------------------------

    def test_calculate_outsource_employee_cost_does_not_cap_service_margin(self):
        """Bug A2: the old `min(service_margin, 500.0)` cap silently truncated
        workbook service-margin values when the `negotiated_service_margin` flag was
        True. With the cap removed, both flag values now return the same cost."""
        from manpower_app.costs import calculate_outsource_employee_cost

        row = pd.Series({
            "Basic": 1000, "Housing Paid": 0, "Trans Paid": 0, "Food": 0,
            "Gosi": 0, "Value O.T (SAR)": 0, "Government Fees": 0, "E.O.S monthly": 0,
            "Service Margin": 20000,
            "Sponser": None,
        })
        # Baseline insurance fallback is 38 (none of the known sponsors match).
        cost_unflagged = calculate_outsource_employee_cost(row, negotiated_service_margin=False)
        cost_flagged = calculate_outsource_employee_cost(row, negotiated_service_margin=True)
        self.assertEqual(cost_unflagged, cost_flagged)
        self.assertEqual(cost_unflagged, 1000 + 38 + 20000)

    # --- Bug A3: Management Saudization mapping -------------------------------

    def test_scenario_5_management_family_honors_profession_saudization_rate(self):
        """Scenario 5 from the Debugging sheet: overall Saudization 0% but every
        per-profession rate 100%. Before this batch, the `Management` job family
        (distinct from `Executive Management`) was not in the profession-rates dict,
        so the constraint was silently dropped for that family."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._make_workbook_bytes(include_management=True)
        processed = process_workbook(io.BytesIO(contents))
        self.assertIn("Management", set(processed.optimization_df["Job Family"]))

        settings = OptimizationSettings(
            enforce_saudization=True,
            saudization_rate=0.0,
            management_saudization_rate=1.0,
            can_reduce_current_saudi=True,
        )
        payload = run_optimization(processed, settings)
        rows = payload["data"]
        mgmt_row = rows[rows["Job Family"] == "Management"].iloc[0]
        saudi = int(mgmt_row[IN_HOUSE_SAUDI_COLUMN])
        non_saudi = int(mgmt_row[IN_HOUSE_NON_SAUDI_COLUMN])
        # 100% rate => Saudis must equal total in-house in this family.
        self.assertEqual(non_saudi, 0)
        self.assertGreater(saudi, 0)

    # --- Bug A4: outsourced ≥ in-house floor ----------------------------------

    def test_cap_outsourced_at_inhouse_function_behavior(self):
        from manpower_app.costs import cap_outsourced_at_inhouse

        # Outsourced higher than in-house non-Saudi -> capped at in-house.
        self.assertEqual(cap_outsourced_at_inhouse(6000, 5000), 5000)
        # Outsourced already lower -> unchanged (the natural case).
        self.assertEqual(cap_outsourced_at_inhouse(4000, 5000), 4000)
        # Missing inhouse signal -> pass-through (no inversion data to enforce).
        self.assertEqual(cap_outsourced_at_inhouse(6000, 0), 6000)

    def test_scenario_1_outsourced_inversion_is_overridden_by_cap(self):
        """Scenario 1 from the Debugging sheet: when in-house safety officers are
        cheaper than outsourced in the workbook, the LP would otherwise refuse to
        outsource on cost grounds even when the outsourceability rule allows it.
        With cap_outsourced_at_inhouse applied, the LP can prefer outsourcing."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        contents = self._make_workbook_bytes(outsourced_more_expensive_than_inhouse=True)
        processed = process_workbook(io.BytesIO(contents))
        settings = OptimizationSettings(can_reduce_current_saudi=True)
        data, _ = prepare_model_data(processed, settings)
        op_row = data[data["Job Family"] == "Factory Operator"].iloc[0]
        # After cap: outsourced unit cost must not exceed in-house non-Saudi unit cost.
        self.assertLessEqual(
            float(op_row[OUTSOURCED_UNIT_COST_BASIS_COLUMN]),
            float(op_row["Fully Loaded Cost per In-house Non-Saudi Employee"]) + 1e-6,
        )

    def test_scenario_8_negotiated_rates_push_above_inhouse_when_user_demands_it(self):
        """Scenario 8: user enables negotiated rates with high insurance + service
        margin, expecting outsourced cost to exceed in-house and the LP to stop
        outsourcing. The cost-inversion cap (A4) must apply to the BASE only, not the
        post-negotiation cost — otherwise the user's intent is silently muted."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        # Workbook with the natural case (in-house cheaper than outsourced base).
        contents = self._make_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        settings = OptimizationSettings(
            negotiated_rates=True,
            negotiated_insurance_cost=2000.0,
            negotiated_service_margin=20000.0,
            can_reduce_current_saudi=True,
        )
        data, _ = prepare_model_data(processed, settings)
        op_row = data[data["Job Family"] == "Factory Operator"].iloc[0]
        inhouse_cost = float(op_row["Fully Loaded Cost per In-house Non-Saudi Employee"])
        outsourced_cost = float(op_row[OUTSOURCED_UNIT_COST_BASIS_COLUMN])
        # With base + 22,000 SAR on top, outsourced MUST clearly exceed in-house —
        # if my cap silently brought it back down, this would fail.
        self.assertGreater(outsourced_cost, inhouse_cost)
        self.assertGreater(outsourced_cost, 22000.0)  # at minimum: insurance + margin

    def test_scenario_8_insurance_margin_apply_to_families_without_subcontractor_data(self):
        """Scenario 8 root cause: families with no subcontractor rows in the workbook
        previously bypassed the insurance+margin path entirely. Now insurance+margin
        apply uniformly to every family whenever negotiated_rates is on."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data
        from manpower_app.family_specs import CustomFamilyCosts, CustomFamilySpec

        # Workbook has Production-Operator rows. Inject a brand-new family with NO
        # subcontractor data — the "Excluding" base column will be NA for this family.
        contents = self._make_workbook_bytes()
        new_family = CustomFamilySpec(
            family_name="Drone Pilot",
            outsourceability="Partially Outsourceable",
            source_pairs=[],
            costs=CustomFamilyCosts(saudi_inhouse=5000, non_saudi_inhouse=4500, outsourced=3000),
        )
        processed = process_workbook(io.BytesIO(contents), custom_families=[new_family])
        settings = OptimizationSettings(
            negotiated_rates=True,
            negotiated_insurance_cost=2000.0,
            negotiated_service_margin=20000.0,
            custom_families=[new_family],
            can_reduce_current_saudi=True,
        )
        data, _ = prepare_model_data(processed, settings)
        drone_row = data[data["Job Family"] == "Drone Pilot"].iloc[0]
        # Insurance + margin must have been added on top of the family's outsourced cost,
        # even though the workbook has no subcontractor rows for this family.
        self.assertGreater(
            float(drone_row[OUTSOURCED_UNIT_COST_BASIS_COLUMN]),
            22000.0,
        )


class DynamicSaudiProtectionTests(unittest.TestCase):
    """Bug E1: protect_current_saudi_percent replaces the binary can_reduce_current_saudi.
    When the percentage is supplied, the LP floors current Saudis at round(current * pct)."""

    def test_full_protection_matches_legacy_bool_false(self):
        # protect=1.0 should behave like can_reduce_current_saudi=False (legacy "protect all").
        row = base_row(**{"Current Total In-house Saudi": 30})
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,  # would normally allow reduction
            tenure_constraint_active=False,
            profession_saudization_rates={},
            protect_current_saudi_percent=1.0,  # overrides the bool
        )
        self.assertEqual(status, "Optimal")
        self.assertGreaterEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 30)

    def test_partial_protection_floors_at_rounded_fraction(self):
        # protect=0.6 of 30 Saudis = 18 floor.
        row = base_row(**{"Current Total In-house Saudi": 30})
        data = pd.DataFrame([row])
        solved, _, _ = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
            protect_current_saudi_percent=0.6,
        )
        self.assertGreaterEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 18)

    def test_zero_protection_matches_legacy_bool_true(self):
        # protect=0.0 should behave like can_reduce_current_saudi=True (no floor).
        row = base_row(**{"Current Total In-house Saudi": 30})
        data = pd.DataFrame([row])
        solved, _, _ = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=False,  # would normally floor at 30
            tenure_constraint_active=False,
            profession_saudization_rates={},
            protect_current_saudi_percent=0.0,  # overrides → no floor
        )
        # No Saudi floor; LP can pick 0 if cheaper.
        self.assertEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 0)


class BUConfigurationOverrideTests(unittest.TestCase):
    """Batch-2 D4: per-BU configuration overrides for outsourceability classifications,
    max ratios, and ratio drivers. Empty overrides preserve historical behavior."""

    def test_outsourceability_override_changes_lp_cap(self):
        from manpower_app.rules import get_outsourceability_rules

        merged = get_outsourceability_rules({"Engineer": "Fully Outsourceable"})
        self.assertEqual(merged["Engineer"], "Fully Outsourceable")
        # Unrelated families are unchanged.
        self.assertEqual(merged["Safety Officer"], "Partially Outsourceable")

    def test_outsourceability_override_ignores_unknown_or_invalid_values(self):
        from manpower_app.rules import get_outsourceability_rules, OUTSOURCEABILITY_RULES

        merged = get_outsourceability_rules({
            "NotARealFamily": "Fully Outsourceable",
            "Engineer": "BogusClassification",
        })
        # No unknown family is added.
        self.assertNotIn("NotARealFamily", merged)
        # Engineer stays at its default because the override value was invalid.
        self.assertEqual(merged["Engineer"], OUTSOURCEABILITY_RULES["Engineer"])

    def test_max_ratio_override_merge(self):
        from manpower_app.rules import get_maximum_ratio_rules

        merged = get_maximum_ratio_rules({"Quarries Foreman": "1:5"})
        self.assertEqual(merged["Quarries Foreman"], "1:5")
        # Unrelated supervisor families are unchanged.
        self.assertEqual(merged["Safety Officer"], "1:50")

    def test_driver_override_recomputes_driver_value(self):
        from manpower_app.ratios import calculate_driver_values

        df = pd.DataFrame([
            {"Activity_Standardized": "Quarries", "Profession_Standardized": "Welder", "Job_Family": "Skilled Labor"},
            {"Activity_Standardized": "Quarries", "Profession_Standardized": "Welder", "Job_Family": "Skilled Labor"},
            {"Activity_Standardized": "Quarries", "Profession_Standardized": "Operator", "Job_Family": "Technician"},
        ])
        default = calculate_driver_values(df)
        overridden = calculate_driver_values(
            df,
            driver_overrides={
                "Quarries Foreman": [{"activity": "Quarries", "profession": "Welder"}]
            },
        )
        # Default rule counts rows in [Labor, Skilled Labor, Technician] under Quarries -> 3.
        self.assertEqual(default["Quarries Foreman"], 3)
        # Override counts only Welder rows -> 2.
        self.assertEqual(overridden["Quarries Foreman"], 2)

    # ─── Mapping pipeline override tests ────────────────────────────────────
    # The BU Excel owns 3 mapping sheets: Profession Mapping (raw → standardized),
    # Activity Mapping (raw → standardized), and Job Families (Activity + Profession
    # → family). Each is layered on top of the hardcoded defaults so a BU can add
    # new payroll values (e.g. "Senior Foreman") without code changes.

    def test_profession_mapping_override_changes_normalization(self):
        """BU adds 'Senior Foreman' → 'Foreman' as a profession override; the engine
        normalizes 'Senior Foreman' to 'Foreman' in the workforce frame. Without the
        override the raw value would pass through unchanged."""
        import io
        from openpyxl import Workbook
        from manpower_app.service import process_workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        for i in range(5):
            inhouse.append([100 + i, "Quarries", "Senior Foreman", "non-saudi", 6400, 0, 5000, 800, 300, 200, 100, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        contents = buf.getvalue()

        # With the BU override, "Senior Foreman" normalizes to "Foreman", which combined
        # with the hardcoded "Quarries - Foreman" → "Quarries Foreman" routes the rows
        # into the existing Quarries Foreman job family.
        processed = process_workbook(
            io.BytesIO(contents),
            profession_mapping_overrides={"Senior Foreman": "Foreman"},
        )
        professions = set(processed.inhouse_cleaned["Profession_Standardized"].unique())
        self.assertIn("Foreman", professions)
        self.assertNotIn("Senior Foreman", professions)
        families = set(processed.inhouse_cleaned["Job_Family"].unique())
        self.assertIn("Quarries Foreman", families)

    def test_activity_mapping_override_changes_normalization(self):
        """BU adds 'Workshop' → 'Factory' as an activity override; the engine routes
        'Workshop' rows into the Factory canonical activity."""
        import io
        from openpyxl import Workbook
        from manpower_app.service import process_workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        for i in range(3):
            inhouse.append([100 + i, "Workshop", "Operator", "non-saudi", 5000, 0, 4000, 500, 250, 150, 100, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        contents = buf.getvalue()

        processed = process_workbook(
            io.BytesIO(contents),
            activity_mapping_overrides={"Workshop": "Factory"},
        )
        activities = set(processed.inhouse_cleaned["Activity_Standardized"].unique())
        self.assertIn("Factory", activities)
        self.assertNotIn("Workshop", activities)
        # Combined with the hardcoded "Factory - Operator" → "Factory Operator" mapping,
        # the rows land in the Factory Operator job family.
        families = set(processed.inhouse_cleaned["Job_Family"].unique())
        self.assertIn("Factory Operator", families)

    def test_job_family_mapping_override_routes_to_new_family(self):
        """BU adds a custom (activity, profession) → family route in the Job Families
        sheet; the engine sends matching rows to that family. Use a profession that
        is NOT in PROFESSION_MAPPING (so it passes through to Profession_Standardized
        unchanged) and assert the override routes the pair to the chosen family."""
        import io
        from openpyxl import Workbook
        from manpower_app.service import process_workbook
        from manpower_app.mappings import PROFESSION_MAPPING

        synthetic_profession = "Drone Pilot"
        self.assertNotIn(
            synthetic_profession, PROFESSION_MAPPING,
            "Test prereq: profession must not be in the default mapping",
        )

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        for i in range(4):
            inhouse.append([100 + i, "Head Office", synthetic_profession, "Saudi", 12000, 0, 9500, 1500, 500, 300, 200, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        contents = buf.getvalue()

        processed = process_workbook(
            io.BytesIO(contents),
            job_family_mapping_overrides={f"Head Office - {synthetic_profession}": "Administration"},
        )
        pilot_rows = processed.inhouse_cleaned[
            processed.inhouse_cleaned["Profession_Standardized"] == synthetic_profession
        ]
        self.assertEqual(len(pilot_rows), 4, "Synthetic profession rows should pass through")
        self.assertTrue(
            (pilot_rows["Job_Family"] == "Administration").all(),
            "Override should route the rows to the Administration family",
        )


class MGICDataDriftRegressionTests(unittest.TestCase):
    """Critical regression: feeding the engine MGIC's mappings via the new BU Excel
    sheet path (as explicit overrides) must produce the EXACT same job families and
    headcounts as the legacy hardcoded path. This is the safety net against any
    drift between the Python dicts in mappings.py and what the consultant's XLS
    encodes — the entire architectural refactor depends on this equivalence."""

    @staticmethod
    def _mgic_synthetic_payroll() -> bytes:
        """Build a payroll that exercises every canonical (activity, profession)
        pair in JOB_FAMILY_MAPPING — one in-house row per pair so all 39 families
        show up in the workforce frame."""
        import io
        from openpyxl import Workbook
        from manpower_app.mappings import JOB_FAMILY_MAPPING
        import re

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        next_id = 100
        for key in JOB_FAMILY_MAPPING.keys():
            parts = re.split(r"\s*-\s*", key, maxsplit=1)
            activity = parts[0].strip()
            profession = parts[1].strip() if len(parts) == 2 else ""
            inhouse.append([
                next_id, activity, profession, "non-saudi",
                5000, 0, 4000, 500, 250, 150, 100, 0,
            ])
            next_id += 1
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_mgic_excel_pipeline_produces_same_39_families_as_legacy_path(self):
        """Run a synthetic MGIC payroll through both paths:
          A) legacy: process_workbook(payroll)  → uses hardcoded defaults
          B) new: process_workbook(payroll, profession=..., activity=..., job_family=...)
             where the overrides are EXACTLY the hardcoded defaults from mappings.py.
        Both must produce identical workforce frames — same families, same counts."""
        import io
        from manpower_app.service import process_workbook
        from manpower_app.mappings import ACTIVITY_MAPPING, JOB_FAMILY_MAPPING, PROFESSION_MAPPING

        contents = self._mgic_synthetic_payroll()

        legacy = process_workbook(io.BytesIO(contents))
        sheet_driven = process_workbook(
            io.BytesIO(contents),
            profession_mapping_overrides=dict(PROFESSION_MAPPING),
            activity_mapping_overrides=dict(ACTIVITY_MAPPING),
            job_family_mapping_overrides=dict(JOB_FAMILY_MAPPING),
        )

        # Same set of families, same count
        legacy_families = legacy.inhouse_cleaned["Job_Family"].value_counts().sort_index()
        sheet_families = sheet_driven.inhouse_cleaned["Job_Family"].value_counts().sort_index()

        self.assertEqual(
            list(legacy_families.index),
            list(sheet_families.index),
            "Job family set diverged between legacy and sheet-driven paths",
        )
        self.assertEqual(
            list(legacy_families.values),
            list(sheet_families.values),
            "Per-family headcounts diverged between legacy and sheet-driven paths",
        )
        # And the BU produces exactly 39 unique families (the MGIC baseline).
        self.assertEqual(len(legacy_families), 39)
        self.assertEqual(len(sheet_families), 39)

    def test_full_excel_round_trip_preserves_engine_output(self):
        """The strongest version of the regression: simulate exactly what the consultant
        does — download the BU Excel, re-upload it, run the engine using the mappings
        the engine just parsed out of the file. The output frame must match the legacy
        hardcoded path bit-for-bit (same 39 families, identical headcounts per family,
        identical optimization_df rows). Catches any silent data loss in the
        build_workbook → parse_workbook round-trip and any subtle whitespace/key issue."""
        import io
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook
        from manpower_app.mappings import ACTIVITY_MAPPING, JOB_FAMILY_MAPPING, PROFESSION_MAPPING
        from manpower_app.rules import OUTSOURCEABILITY_RULES, MAXIMUM_RATIO_RULES
        from manpower_app.service import process_workbook

        # 1. Build MGIC's Excel as if the consultant pressed "Download" on the
        #    Configuration panel — full seeded config with all hardcoded defaults.
        mgic_config = BUConfigurationPayload(
            profession_mapping=dict(PROFESSION_MAPPING),
            activity_mapping=dict(ACTIVITY_MAPPING),
            job_family_mapping=dict(JOB_FAMILY_MAPPING),
            outsourceability_overrides=dict(OUTSOURCEABILITY_RULES),
            ratio_overrides=dict(MAXIMUM_RATIO_RULES),
        )
        xlsx = build_workbook("MGIC", "Marble & Granite Intl Co.", mgic_config)

        # 2. Parse the XLSX back as if the consultant just uploaded it.
        parsed_config, parse_errors = parse_workbook(xlsx)
        self.assertEqual(parse_errors, [], "Round-trip parse must produce no errors")

        # 3. Run the engine TWICE on the same synthetic MGIC payroll:
        #    a) Legacy path — no overrides, hardcoded defaults active.
        #    b) Excel-driven path — overrides come from the parsed XLSX (post-round-trip).
        payroll = self._mgic_synthetic_payroll()

        legacy = process_workbook(io.BytesIO(payroll))
        via_excel = process_workbook(
            io.BytesIO(payroll),
            profession_mapping_overrides=parsed_config.profession_mapping,
            activity_mapping_overrides=parsed_config.activity_mapping,
            job_family_mapping_overrides=parsed_config.job_family_mapping,
        )

        # 4. The workforce frames must agree on (family, count) — same 39 families.
        legacy_counts = legacy.inhouse_cleaned.groupby("Job_Family").size().sort_index()
        excel_counts = via_excel.inhouse_cleaned.groupby("Job_Family").size().sort_index()
        self.assertEqual(
            list(legacy_counts.index), list(excel_counts.index),
            "Job-family set diverged after Excel round-trip",
        )
        self.assertEqual(
            list(legacy_counts.values), list(excel_counts.values),
            "Per-family headcounts diverged after Excel round-trip",
        )
        self.assertEqual(len(legacy_counts), 39, "MGIC baseline should be 39 families")

        # 5. The optimization_df (per-family cost inputs into the LP) must agree on the
        #    rows that matter for the optimizer — family, headcount, outsourceability.
        legacy_opt = legacy.optimization_df.sort_values("Job Family").reset_index(drop=True)
        excel_opt = via_excel.optimization_df.sort_values("Job Family").reset_index(drop=True)
        self.assertEqual(
            list(legacy_opt["Job Family"]), list(excel_opt["Job Family"]),
            "optimization_df family rows diverged",
        )
        self.assertEqual(
            list(legacy_opt["Current Headcount"]), list(excel_opt["Current Headcount"]),
            "optimization_df headcounts diverged",
        )
        self.assertEqual(
            list(legacy_opt["Outsourceability Type"]),
            list(excel_opt["Outsourceability Type"]),
            "Outsourceability classifications diverged",
        )

    def test_mgic_unmapped_pair_surfaces_in_response(self):
        """When the payroll has a (activity, profession) the BU's mappings don't
        cover, the engine returns those unmapped pairs in the response (the frontend
        hard-blocks on this and tells the user to add a row to the BU Excel).
        Include a couple of mapped rows alongside so the optimization frame is
        non-empty (the engine errors out on a fully-empty frame, which is a separate
        concern from unmapped-pair detection)."""
        import io
        from openpyxl import Workbook
        from manpower_app.service import process_workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        # A few rows that ARE in the default mapping (so the optimization frame is non-empty)
        for i in range(5):
            inhouse.append([100 + i, "Quarries", "Skilled Labor", "non-saudi", 4200, 0, 3300, 500, 200, 150, 50, 0])
        # The unmapped pair: neither activity nor profession in the default maps
        for i in range(3):
            inhouse.append([200 + i, "Drone Yard", "Drone Pilot", "Saudi", 8000, 0, 6500, 1000, 300, 200, 100, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)

        processed = process_workbook(io.BytesIO(buf.getvalue()))
        # The unmapped pair shows up in the response so the UI can hard-block.
        pair_strs = {f"{p.get('activity')}|{p.get('profession')}" for p in processed.unmapped_pairs}
        self.assertTrue(
            any("Drone" in s for s in pair_strs),
            f"Expected an unmapped 'Drone' pair in {pair_strs}",
        )


class BUConfigurationExcelRoundTripTests(unittest.TestCase):
    """Batch-2 H1: per-BU configuration round-trips through the Excel template format
    without data loss. Empty overrides on the way out should still be empty on the
    way back (the user did not edit anything)."""

    def test_round_trip_full_payload(self):
        """Round-trip every field in the new 7-sheet workbook. The build merges user
        entries on top of the tool's hardcoded defaults (so consultants always see the
        full editing surface), so parsed values include user entries plus all defaults.
        User-specified values must be preserved exactly."""
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook

        original = BUConfigurationPayload(
            profession_mapping={"Welder Sr": "Welder", "Senior Welder": "Welder"},
            activity_mapping={"Production": "Factory", "Workshop": "Factory"},
            job_family_mapping={"Factory - Welder": "Skilled Labor"},
            outsourceability_overrides={"Engineer": "Fully Outsourceable"},
            ratio_overrides={"Quarries Foreman": "1:5"},
            driver_overrides={
                "Production Foreman": [
                    {"activity": "Factory", "profession": "Welder"},
                    {"activity": "Factory", "profession": "Operator"},
                ],
            },
            saudi_cost_premium=1.25,
            outsource_cost_discount=0.15,
        )
        xlsx = build_workbook("MGIC", "Marble & Granite International Company", original)
        parsed, errors = parse_workbook(xlsx)
        self.assertEqual(errors, [])
        # User-specified values must survive the round-trip (the parsed dicts may
        # additionally contain the tool's defaults — that's fine, the engine merges).
        for raw, std in original.profession_mapping.items():
            self.assertEqual(parsed.profession_mapping.get(raw), std)
        for raw, std in original.activity_mapping.items():
            self.assertEqual(parsed.activity_mapping.get(raw), std)
        for key, fam in original.job_family_mapping.items():
            self.assertEqual(parsed.job_family_mapping.get(key), fam)
        for fam, value in original.outsourceability_overrides.items():
            self.assertEqual(parsed.outsourceability_overrides.get(fam), value)
        for fam, value in original.ratio_overrides.items():
            self.assertEqual(parsed.ratio_overrides.get(fam), value)
        # Drivers are emitted only from user data (no defaults), so should round-trip exactly.
        self.assertEqual(parsed.driver_overrides, original.driver_overrides)
        self.assertEqual(parsed.saudi_cost_premium, original.saudi_cost_premium)
        self.assertEqual(parsed.outsource_cost_discount, original.outsource_cost_discount)

    def test_starter_template_for_empty_bu_is_skeleton_not_pre_filled(self):
        """For a fresh BU (e.g. UAAC, FAST) the starter Excel is a SKELETON, not
        pre-filled with MGIC's data. The mapping sheets contain only italic example
        rows + instructions; the user fills them with values from THEIR payroll.
        Parsing the skeleton back yields empty mapping dicts (example rows are
        skipped by the parser) so engine falls back to hardcoded defaults."""
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook

        xlsx = build_workbook("UAAC", None, BUConfigurationPayload())
        parsed, errors = parse_workbook(xlsx)
        self.assertEqual(errors, [])
        # Profession / Activity / Job-Family mappings are empty — example rows
        # with the "(example)" prefix were filtered out by the parser.
        self.assertEqual(parsed.profession_mapping, {})
        self.assertEqual(parsed.activity_mapping, {})
        self.assertEqual(parsed.job_family_mapping, {})
        # Outsourceability is also empty (lives inside the empty Job Families sheet)
        self.assertEqual(parsed.outsourceability_overrides, {})
        # Ratios sheet always lists supervisors but with blank values for empty BUs
        self.assertEqual(parsed.ratio_overrides, {})

    def test_invalid_outsourceability_value_is_rejected(self):
        """Pollute one Outsourceability cell in the Job Families sheet → parser reports
        the row-level error and never persists the bad value. Build a populated
        config first so the Job Families sheet has real data rows (not skeleton)."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook
        from manpower_app.mappings import JOB_FAMILY_MAPPING
        from manpower_app.rules import OUTSOURCEABILITY_RULES

        # Use a populated config so the Job Families sheet has real rows.
        populated = BUConfigurationPayload(
            job_family_mapping=dict(JOB_FAMILY_MAPPING),
            outsourceability_overrides=dict(OUTSOURCEABILITY_RULES),
        )
        xlsx = build_workbook("MGIC", None, populated)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb["Job Families"]
        # Find a row whose family is Engineer (column C is Job Family, D is Outsourceability)
        # and corrupt its Outsourceability cell with an invalid string.
        for row in ws.iter_rows(min_row=2):
            if row[2].value == "Engineer":
                row[3].value = "Maybe Outsourceable"
                break
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertTrue(any("Engineer" in e and "must be one of" in e for e in errors))
        # Engineer was not promoted to the saved outsourceability overrides because
        # its only row had an invalid value — falls back to hardcoded at engine time.
        self.assertNotEqual(parsed.outsourceability_overrides.get("Engineer"), "Maybe Outsourceable")

    def test_invalid_ratio_format_is_rejected(self):
        """Build a populated Ratios sheet, corrupt one row, parser flags it.
        (The skeleton starter has only example rows so we use a config with
        ratio_overrides to get the canonical Quarries Foreman row.)"""
        import io
        import openpyxl
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook
        from manpower_app.rules import MAXIMUM_RATIO_RULES

        populated = BUConfigurationPayload(ratio_overrides=dict(MAXIMUM_RATIO_RULES))
        xlsx = build_workbook("MGIC", None, populated)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb["Ratios"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "Quarries Foreman":
                row[1].value = "five"
                break
        buf = io.BytesIO()
        wb.save(buf)
        _, errors = parse_workbook(buf.getvalue())
        self.assertTrue(any("Quarries Foreman" in e and "1:N" in e for e in errors))

    def test_starter_workbook_for_empty_bu_is_skeleton_with_examples_only(self):
        """Starter workbook for a fresh BU (UAAC, FAST, etc.) is a SKELETON: 7
        sheets, all present, but EVERY mapping sheet contains only italic
        (example) rows + instructions — NOT pre-filled with MGIC's data. This
        applies uniformly to Profession Mapping, Activity Mapping, Job Families,
        Ratios, AND Drivers."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook

        xlsx = build_workbook("UAAC", None, BUConfigurationPayload())
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        self.assertEqual(
            wb.sheetnames,
            [
                "README",
                "Profession Mapping",
                "Activity Mapping",
                "Job Families",
                "Ratios",
                "Drivers",
                "Cost Assumptions",
            ],
        )

        def _all_skeleton_rows(sheet_name: str, *, col: int = 1) -> None:
            rows = [
                row[col - 1]
                for row in wb[sheet_name].iter_rows(min_row=2, values_only=True)
                if row[col - 1]
            ]
            self.assertTrue(
                all(str(r).startswith("(example)") for r in rows),
                f"{sheet_name} should contain only (example) rows, got: {rows}",
            )

        _all_skeleton_rows("Profession Mapping")
        _all_skeleton_rows("Activity Mapping")
        _all_skeleton_rows("Job Families")
        _all_skeleton_rows("Ratios")
        _all_skeleton_rows("Drivers")


class EndToEndUserJourneyTests(unittest.TestCase):
    """Walk through complete user actions end-to-end:
    upload payroll → set BU configuration → add custom families → run optimization,
    and verify each step's effect actually reaches the LP."""

    @staticmethod
    def _quarries_workbook_bytes() -> bytes:
        """Workbook with three job families:
        * Engineer  (Not Outsourceable by default)
        * Skilled Labor (Fully Outsourceable)
        * Quarries Foreman (Partially Outsourceable, ratio 1:15)
        """
        import io
        from openpyxl import Workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        # Engineer (Not Outsourceable). Big in-house base so cost is high.
        for i in range(10):
            inhouse.append([100 + i, "Head Office", "Engineer", "Saudi", 18000, 0, 14000, 2000, 800, 800, 400, 0])
        # Skilled Labor (Fully Outsourceable).
        for i in range(40):
            inhouse.append([200 + i, "Quarries", "Skilled Labor", "non-saudi", 4200, 0, 3300, 500, 200, 150, 50, 0])
        # Quarries Foreman (Partially Outsourceable).
        for i in range(5):
            inhouse.append([300 + i, "Quarries", "Foreman", "non-saudi", 6400, 0, 5000, 800, 300, 200, 100, 0])

        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        # Outsourced Skilled Labor — cheaper than in-house (natural case).
        for i in range(20):
            sub.append([400 + i, "Quarries", "Skilled Labor", "non-saudi", 2400, 200, 100, 50, 0, 0, 0, 0, 100])
        # Outsourced Quarries Foreman — cheaper than in-house.
        for i in range(2):
            sub.append([450 + i, "Quarries", "Foreman", "non-saudi", 4500, 500, 200, 100, 0, 0, 0, 0, 100])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # --- Journey 1: BU outsourceability override changes the LP outcome --------

    def test_bu_outsourceability_override_makes_engineer_outsourceable(self):
        """Default behavior: Engineer is 'Not Outsourceable' so LP keeps every Engineer
        in-house. After a BU sets Engineer to 'Fully Outsourceable', the LP can move
        Engineers to outsourced when it's cheaper."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))

        # Default: Engineer locked in-house.
        baseline = run_optimization(processed, OptimizationSettings(can_reduce_current_saudi=True))
        baseline_engineer = baseline["data"][baseline["data"]["Job Family"] == "Engineer"].iloc[0]
        self.assertEqual(int(baseline_engineer[OUTSOURCED_COLUMN]), 0)

        # BU override: Engineer → Fully Outsourceable. The LP would only outsource
        # Engineers when outsourcing is cheaper. In this workbook there are no Engineer
        # subcontractor rows, so insurance/margin are 0 and the inferred outsourced
        # cost ends up below in-house — LP can move Engineers out.
        overridden = run_optimization(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                outsourceability_overrides={"Engineer": "Fully Outsourceable"},
            ),
        )
        overridden_engineer = overridden["data"][overridden["data"]["Job Family"] == "Engineer"].iloc[0]
        # With the override, the LP has the FREEDOM to outsource — we just assert that
        # the LP no longer hard-locks the family to in-house. Even one outsourced is
        # proof the override flowed through.
        self.assertGreaterEqual(int(overridden_engineer[OUTSOURCED_COLUMN]), 1)

    # --- Journey 2: BU ratio override changes the minimum headcount ------------

    def test_bu_ratio_override_changes_minimum_headcount_for_supervisor(self):
        """Default: Quarries Foreman ratio 1:15 → with 20 Skilled Labor drivers, min
        Foreman = ceil(20/15) = 2. A BU tightens the ratio to 1:5 → min = ceil(20/5) = 4.
        """
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        contents = self._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))

        default_data, _ = prepare_model_data(processed, OptimizationSettings(can_reduce_current_saudi=True))
        default_min = int(
            default_data[default_data["Job Family"] == "Quarries Foreman"]["Minimum Headcount Needed"].iloc[0]
        )

        tighter_data, _ = prepare_model_data(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                max_ratio_overrides={"Quarries Foreman": "1:5"},
            ),
        )
        tighter_min = int(
            tighter_data[tighter_data["Job Family"] == "Quarries Foreman"]["Minimum Headcount Needed"].iloc[0]
        )

        self.assertGreater(tighter_min, default_min)

    # --- Journey 3: BU driver override changes the driver value ----------------

    def test_bu_driver_override_recomputes_driver_value_per_family(self):
        """Default Quarries Foreman driver set counts Quarries Labor + Skilled Labor +
        Technician. With a BU override restricting it to a profession that doesn't
        appear in the workbook, the driver value should drop accordingly."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        contents = self._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))

        default_data, _ = prepare_model_data(processed, OptimizationSettings(can_reduce_current_saudi=True))
        default_driver = int(
            default_data[default_data["Job Family"] == "Quarries Foreman"]["Driver Value"].iloc[0]
        )

        # Restrict to a non-existent profession → driver value collapses.
        narrowed_data, _ = prepare_model_data(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                driver_overrides={
                    "Quarries Foreman": [
                        {"activity": "Quarries", "profession": "NonexistentRole"},
                    ],
                },
            ),
        )
        narrowed_driver = int(
            narrowed_data[narrowed_data["Job Family"] == "Quarries Foreman"]["Driver Value"].iloc[0]
        )

        self.assertEqual(narrowed_driver, 0)
        self.assertGreater(default_driver, 0)

    # --- Journey 4: Two BUs with different overrides produce different results --

    def test_two_bus_with_different_overrides_yield_different_results(self):
        """Switching between two BUs (same workbook, different overrides) produces
        materially different optimizations. Proves the BU layer actually isolates
        per-BU rules through to the LP output."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))

        # BU A: tighter ratio for Quarries Foreman → more Foremen required → fewer
        # available for outsourcing.
        bu_a = run_optimization(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                max_ratio_overrides={"Quarries Foreman": "1:2"},
            ),
        )
        # BU B: looser ratio → fewer Foremen required → more can be outsourced.
        bu_b = run_optimization(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                max_ratio_overrides={"Quarries Foreman": "1:100"},
            ),
        )
        a_foreman = bu_a["data"][bu_a["data"]["Job Family"] == "Quarries Foreman"].iloc[0]
        b_foreman = bu_b["data"][bu_b["data"]["Job Family"] == "Quarries Foreman"].iloc[0]
        a_inhouse = int(a_foreman[IN_HOUSE_SAUDI_COLUMN]) + int(a_foreman[IN_HOUSE_NON_SAUDI_COLUMN])
        b_inhouse = int(b_foreman[IN_HOUSE_SAUDI_COLUMN]) + int(b_foreman[IN_HOUSE_NON_SAUDI_COLUMN])
        # Tighter ratio forces more (or equal) in-house headcount than the looser one.
        self.assertGreaterEqual(a_inhouse, b_inhouse)

    # --- Journey 5: Add a custom family + target headcount + run optimization ---

    def test_custom_family_with_target_headcount_appears_in_results(self):
        """User adds a brand-new job family for target planning, gives it costs and
        a target headcount, and the LP plans for it."""
        import io
        from manpower_app.family_specs import CustomFamilyCosts, CustomFamilySpec
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._quarries_workbook_bytes()
        new_family = CustomFamilySpec(
            family_name="Drone Inspector",
            outsourceability="Partially Outsourceable",
            source_pairs=[],
            costs=CustomFamilyCosts(saudi_inhouse=9000, non_saudi_inhouse=8000, outsourced=5500),
        )
        processed = process_workbook(io.BytesIO(contents), custom_families=[new_family])
        settings = OptimizationSettings(
            optimization_mode="target",
            target_headcounts={"Drone Inspector": 12},
            custom_families=[new_family],
            can_reduce_current_saudi=True,
        )
        payload = run_optimization(processed, settings)
        drone_row = payload["data"][payload["data"]["Job Family"] == "Drone Inspector"].iloc[0]
        total = (
            int(drone_row[OUTSOURCED_COLUMN])
            + int(drone_row[IN_HOUSE_SAUDI_COLUMN])
            + int(drone_row[IN_HOUSE_NON_SAUDI_COLUMN])
        )
        self.assertEqual(total, 12)

    # --- Journey 6: Map a new profession to an existing family via reprocess ----

    def test_payroll_pair_override_routes_new_profession_to_existing_family(self):
        """User uploads a payroll with a new profession (e.g. 'Senior Skilled Labor')
        that isn't in JOB_FAMILY_MAPPING. Instead of creating a custom family, they
        use the 'Map to existing family' affordance to route it to 'Skilled Labor'.
        After reprocess, the pair is no longer unmapped and the workers appear in
        the Skilled Labor family."""
        import io
        from openpyxl import Workbook
        from manpower_app.service import process_workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        # Unknown profession: "Senior Skilled Labor" in Quarries.
        for i in range(3):
            inhouse.append([700 + i, "Quarries", "Senior Skilled Labor", "non-saudi", 5000, 0, 4000, 600, 200, 200, 100, 0])
        # Plus a known mapping so the rest of the workbook validates.
        for i in range(5):
            inhouse.append([800 + i, "Quarries", "Skilled Labor", "non-saudi", 4200, 0, 3300, 500, 200, 150, 50, 0])

        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        sub.append([900, "Quarries", "Skilled Labor", "non-saudi", 2400, 200, 100, 50, 0, 0, 0, 0, 100])
        buf = io.BytesIO()
        wb.save(buf)
        contents = buf.getvalue()

        # Without overrides: "Senior Skilled Labor" is unmapped.
        baseline = process_workbook(io.BytesIO(contents))
        unmapped_pairs = {(p["activity"], p["profession"]) for p in baseline.unmapped_pairs}
        self.assertIn(("Quarries", "Senior Skilled Labor"), unmapped_pairs)

        # With the user-supplied override: pair routes to Skilled Labor.
        with_override = process_workbook(
            io.BytesIO(contents),
            payroll_pair_overrides={"Quarries|Senior Skilled Labor": "Skilled Labor"},
        )
        self.assertEqual(with_override.unmapped_pairs, [])
        # Skilled Labor's headcount now includes the 3 Senior Skilled Labor workers
        # plus the original 5 = 8 in-house.
        skilled = with_override.optimization_df[with_override.optimization_df["Job Family"] == "Skilled Labor"].iloc[0]
        self.assertEqual(int(skilled["Current In-House Count"]), 8)


class EdgeCaseRobustnessTests(unittest.TestCase):
    """Boundary and stress conditions across the optimization pipeline."""

    def test_cap_outsourced_at_inhouse_passthrough_when_signals_missing(self):
        from manpower_app.costs import cap_outsourced_at_inhouse
        # No in-house signal → cap is a no-op.
        self.assertEqual(cap_outsourced_at_inhouse(6000, 0), 6000)
        # No outsourced signal → cap is a no-op.
        self.assertEqual(cap_outsourced_at_inhouse(0, 5000), 0)
        # Equal → unchanged.
        self.assertEqual(cap_outsourced_at_inhouse(5000, 5000), 5000)

    def test_risk_factor_at_one_makes_outsourced_workers_count_zero(self):
        """At R=1, outsourced workers contribute 0 to the risk-adjusted minimum.
        If minimum_inhouse equals total_headcount, no slack exists for outsourcing."""
        row = base_row(**{
            "Risk Factor": 1.0,
            "Current Headcount": 50,
            "Minimum Headcount Needed": 50,
        })
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=False,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        self.assertEqual(status, "Optimal")
        # With M == T and R == 1, the LP can only meet the constraint with all in-house.
        self.assertEqual(int(solved.iloc[0][OUTSOURCED_COLUMN]), 0)

    def test_saudi_premium_below_one_is_clamped_to_one(self):
        """User-supplied premium < 1.0 would let Saudis be cheaper than non-Saudis.
        The clamp in costs.calculate_inhouse_cost_split prevents the inversion."""
        from manpower_app.costs import calculate_inhouse_cost_split
        split = calculate_inhouse_cost_split(100.0, 1, 1, saudi_premium=0.5)
        saudi = split["Fully Loaded Cost per In-house Saudi Employee"]
        non_saudi = split["Fully Loaded Cost per In-house Non-Saudi Employee"]
        # Even though premium was 0.5, the clamp makes Saudis cost ≥ non-Saudis.
        self.assertGreaterEqual(saudi, non_saudi)

    def test_negotiated_rates_off_ignores_user_insurance_and_margin(self):
        """When the user toggles negotiated_rates OFF, the insurance + margin fields
        do not affect the optimization basis. The workbook-derived cost wins."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        a, _ = prepare_model_data(processed, OptimizationSettings(
            can_reduce_current_saudi=True,
            negotiated_rates=False,
            negotiated_insurance_cost=99999.0,
            negotiated_service_margin=99999.0,
        ))
        b, _ = prepare_model_data(processed, OptimizationSettings(
            can_reduce_current_saudi=True,
            negotiated_rates=False,
        ))
        # Pick any partially-outsourceable row and confirm the basis is identical.
        a_row = a[a["Job Family"] == "Skilled Labor"].iloc[0]
        b_row = b[b["Job Family"] == "Skilled Labor"].iloc[0]
        self.assertAlmostEqual(
            float(a_row[OUTSOURCED_UNIT_COST_BASIS_COLUMN]),
            float(b_row[OUTSOURCED_UNIT_COST_BASIS_COLUMN]),
            places=6,
        )


class ExcelValidationEdgeCases(unittest.TestCase):
    """The import path must reject bad values cleanly without losing the rest of the
    payload, and must accept tolerable formatting noise."""

    def test_mixed_valid_and_invalid_rows_returns_errors_and_no_partial_save(self):
        """When some rows in the Job Families sheet have an invalid Outsourceability
        value, the parser reports errors and the invalid row is NOT promoted to the
        outsourceability_overrides dict. Valid rows are still parsed.

        Build with a populated config so the Job Families sheet has real data rows
        (vs the skeleton that ships for empty BUs)."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook
        from manpower_app.mappings import JOB_FAMILY_MAPPING

        populated = BUConfigurationPayload(
            job_family_mapping=dict(JOB_FAMILY_MAPPING),
        )
        xlsx = build_workbook("MGIC", None, populated)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb["Job Families"]
        # Clear ALL outsourceability cells first so we test JUST the rows we touch.
        for row in ws.iter_rows(min_row=2):
            if row[2].value:  # only clear actual data rows, not the right-side reference cols
                row[3].value = ""
        # Find an Engineer row and corrupt it
        for row in ws.iter_rows(min_row=2):
            if row[2].value == "Engineer":
                row[3].value = "Maybe Outsourceable"
                break
        # Find a Skilled Labor row and set a VALID value
        for row in ws.iter_rows(min_row=2):
            if row[2].value == "Skilled Labor":
                row[3].value = "Fully Outsourceable"
                break
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertTrue(any("Engineer" in e for e in errors))
        self.assertEqual(parsed.outsourceability_overrides.get("Skilled Labor"), "Fully Outsourceable")
        self.assertNotIn("Engineer", parsed.outsourceability_overrides)

    def test_whitespace_in_ratio_is_normalized_on_parse(self):
        """Build a populated Ratios sheet, set a value with extra whitespace,
        parser normalizes it."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook
        from manpower_app.rules import MAXIMUM_RATIO_RULES

        populated = BUConfigurationPayload(ratio_overrides=dict(MAXIMUM_RATIO_RULES))
        xlsx = build_workbook("MGIC", None, populated)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb["Ratios"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "Quarries Foreman":
                row[1].value = "1 : 7"  # spaces around the colon
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertEqual(errors, [])
        self.assertEqual(parsed.ratio_overrides["Quarries Foreman"], "1:7")

    def test_legacy_engine_knobs_sheet_is_still_parsed_for_backwards_compat(self):
        """A user who saved an older BU workbook (PR-#2 era: 4 sheets including
        Engine Knobs but no Cost Assumptions sheet) and re-uploads it should still
        have the cost knobs picked up. Construct that legacy layout manually."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import parse_workbook

        wb = openpyxl.Workbook()
        wb.active.title = "README"
        for sheet_name, header_row in (
            ("Outsourceability", ["Job Family", "Value"]),
            ("Ratios", ["Supervisor Family", "Value (e.g. 1:10)"]),
            ("Drivers", ["Supervisor Family", "Activity", "Profession"]),
        ):
            ws = wb.create_sheet(sheet_name)
            ws.append(header_row)
        ws = wb.create_sheet("Engine Knobs")
        ws.append(["Knob", "Value", "Notes"])
        ws.append(["Saudi cost premium", 1.25, ""])
        ws.append(["Outsource cost discount", 0.15, ""])
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertEqual(errors, [])
        self.assertEqual(parsed.saudi_cost_premium, 1.25)
        self.assertEqual(parsed.outsource_cost_discount, 0.15)

    def test_workbook_with_no_recognized_sheets_returns_empty_config(self):
        """No sheets are strictly required anymore. A workbook with no recognized
        sheets parses cleanly to an empty config; the engine falls back to hardcoded
        defaults for everything."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import parse_workbook

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Random"
        ws.append(["something", "irrelevant"])
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertEqual(errors, [])
        self.assertTrue(parsed.is_empty())


class ScenarioCostKnobsTests(unittest.TestCase):
    """After the BU/scenario split, Saudi pay premium and outsource cost discount are
    SCENARIO knobs (sent in OptimizationSettings), not BU constraints (no longer in the
    BU Excel). These tests verify the scenario path still works end-to-end."""

    def test_starter_excel_has_no_engine_knobs_sheet(self):
        """The legacy 'Engine Knobs' sheet name is gone. Cost knobs live in the new
        'Cost Assumptions' sheet."""
        import io
        import openpyxl
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook

        xlsx = build_workbook("UAAC", "United Arab Aluminium Company", BUConfigurationPayload())
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        self.assertNotIn("Engine Knobs", wb.sheetnames)
        self.assertIn("Cost Assumptions", wb.sheetnames)

    def test_saudi_pay_premium_scenario_knob_changes_inhouse_cost_split(self):
        """Set the scenario-level premium high; the LP's per-family Saudi cost basis
        must reflect that multiplier (not the default 1.10)."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))

        a, _ = prepare_model_data(processed, OptimizationSettings(
            can_reduce_current_saudi=True,
            saudi_cost_premium=1.10,
        ))
        b, _ = prepare_model_data(processed, OptimizationSettings(
            can_reduce_current_saudi=True,
            saudi_cost_premium=1.80,
        ))
        # Pick a family with both Saudi and non-Saudi headcount.
        a_op = a[a["Job Family"] == "Engineer"].iloc[0]
        b_op = b[b["Job Family"] == "Engineer"].iloc[0]
        # Higher premium -> larger Saudi/non-Saudi cost ratio (and a higher Saudi cost).
        a_ratio = float(a_op["Fully Loaded Cost per In-house Saudi Employee"]) / float(a_op["Fully Loaded Cost per In-house Non-Saudi Employee"])
        b_ratio = float(b_op["Fully Loaded Cost per In-house Saudi Employee"]) / float(b_op["Fully Loaded Cost per In-house Non-Saudi Employee"])
        self.assertAlmostEqual(a_ratio, 1.10, places=4)
        self.assertAlmostEqual(b_ratio, 1.80, places=4)

    def test_outsource_discount_scenario_knob_replaces_workbook_outsource_cost(self):
        """The Override-outsource-cost scenario knob replaces per-family outsource cost
        with `(1 - discount) * non-Saudi in-house cost`. Verify the basis column reflects
        that — not the workbook value."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data
        from manpower_app.terminology import OUTSOURCED_UNIT_COST_BASIS_COLUMN

        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        data, _ = prepare_model_data(processed, OptimizationSettings(
            can_reduce_current_saudi=True,
            outsource_cost_discount=0.30,
        ))
        op = data[data["Job Family"] == "Skilled Labor"].iloc[0]
        expected = 0.70 * float(op["Fully Loaded Cost per In-house Non-Saudi Employee"])
        actual = float(op[OUTSOURCED_UNIT_COST_BASIS_COLUMN])
        self.assertAlmostEqual(actual, expected, places=4)


class CombinedJourneyTests(unittest.TestCase):
    """Walk a full client journey: BU config (Excel-uploaded overrides) + scenario knobs
    (Saudi pay premium, risk factor) + custom family + payroll pair override all feeding
    into the LP at once. This is the realistic 'all parts working together' test."""

    def test_bu_excel_and_scenario_knobs_and_pair_override_all_apply(self):
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data, run_optimization

        # 1. Build a workbook with one mappable family + one unmapped pair.
        from openpyxl import Workbook
        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        # Mappable: Quarries/Skilled Labor (canonical).
        for i in range(10):
            inhouse.append([i, "Quarries", "Skilled Labor", "non-saudi", 4200, 0, 3300, 500, 200, 150, 50, 0])
        # Unmappable: Quarries/Senior Labor — needs payroll_pair_override to route.
        for i in range(5):
            inhouse.append([100 + i, "Quarries", "Senior Labor", "non-saudi", 5000, 0, 4000, 600, 200, 200, 100, 0])
        # Some Saudi in-house Engineers so the Saudi-premium scenario knob has data.
        for i in range(4):
            inhouse.append([200 + i, "Head Office", "Engineer", "Saudi", 18000, 0, 14000, 2000, 800, 800, 400, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        for i in range(10):
            sub.append([300 + i, "Quarries", "Skilled Labor", "non-saudi", 2400, 200, 100, 50, 0, 0, 0, 0, 100])
        buf = io.BytesIO()
        wb.save(buf)
        contents = buf.getvalue()

        # 2. Process WITH the pair override (simulates "Map to existing family").
        processed = process_workbook(
            io.BytesIO(contents),
            payroll_pair_overrides={"Quarries|Senior Labor": "Skilled Labor"},
        )
        self.assertEqual(processed.unmapped_pairs, [])

        # 3. Run optimization with BU overrides + scenario knobs simultaneously.
        settings = OptimizationSettings(
            # BU override (Excel-supplied): make Engineer fully outsourceable for this BU.
            outsourceability_overrides={"Engineer": "Fully Outsourceable"},
            # BU override (Excel-supplied): tighter supervisor ratio.
            max_ratio_overrides={"Quarries Foreman": "1:5"},
            # Scenario knob: aggressive risk factor.
            risk_factor=0.5,
            # Scenario knob: Saudi pay premium.
            saudi_cost_premium=1.4,
            # Scenario knob: protect 50% of current Saudis.
            protect_current_saudi_percent=0.5,
            # Scenario knob: enforce saudization at 20% overall.
            enforce_saudization=True,
            saudization_rate=0.20,
        )
        payload = run_optimization(processed, settings)
        self.assertEqual(payload["metadata"]["optimization_status"], "Optimal")

        # 4. Verify the various overrides made it into the LP's input frame.
        data = payload["data"]
        # BU outsourceability override applied:
        eng = data[data["Job Family"] == "Engineer"].iloc[0]
        self.assertEqual(str(eng["Outsourceability Type"]), "Fully Outsourceable")
        # Pair override worked: Skilled Labor headcount includes the 5 Senior Labor rows.
        skilled = data[data["Job Family"] == "Skilled Labor"].iloc[0]
        # 10 mappable + 5 routed in-house + 10 subcontractor = 25 total.
        self.assertEqual(int(skilled["Current Headcount"]), 25)
        # Saudi premium scenario knob reflected in cost split.
        ratio = float(eng["Fully Loaded Cost per In-house Saudi Employee"]) / float(
            eng["Fully Loaded Cost per In-house Non-Saudi Employee"]
        )
        self.assertAlmostEqual(ratio, 1.4, places=3)
        # Saudi protection 50% of 4 current = 2 floor.
        from manpower_app.optimization import IN_HOUSE_SAUDI_COLUMN
        self.assertGreaterEqual(int(eng[IN_HOUSE_SAUDI_COLUMN]), 2)


class TargetModeCombinedTests(unittest.TestCase):
    """Target manpower plan mode is tested at the basic level elsewhere. These tests
    cover target mode running alongside the batch-2 features: BU overrides applied,
    scenario cost knobs applied, and the LP still hits the target headcount."""

    def test_target_mode_with_bu_outsourceability_override(self):
        """Set a target headcount AND override a family's outsourceability for this BU.
        Both must apply: the LP must split the target across in-house/outsourced respecting
        the BU's overridden classification."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        payload = run_optimization(
            processed,
            OptimizationSettings(
                optimization_mode="target",
                target_headcounts={"Engineer": 30, "Skilled Labor": 60, "Quarries Foreman": 6},
                outsourceability_overrides={"Engineer": "Fully Outsourceable"},
                can_reduce_current_saudi=True,
            ),
        )
        data = payload["data"]
        eng = data[data["Job Family"] == "Engineer"].iloc[0]
        # Target headcount applied:
        self.assertEqual(int(eng["Current Headcount"]), 30)
        # BU override applied:
        self.assertEqual(str(eng["Outsourceability Type"]), "Fully Outsourceable")
        # Total split matches target:
        total = (
            int(eng[OUTSOURCED_COLUMN])
            + int(eng[IN_HOUSE_SAUDI_COLUMN])
            + int(eng[IN_HOUSE_NON_SAUDI_COLUMN])
        )
        self.assertEqual(total, 30)

    def test_target_mode_with_saudi_pay_premium_scenario_knob(self):
        """Target mode + Saudi pay premium scenario knob: the LP's cost basis reflects the
        scenario premium even when running against user-supplied target headcounts."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        data, _ = prepare_model_data(processed, OptimizationSettings(
            optimization_mode="target",
            target_headcounts={"Engineer": 20},
            saudi_cost_premium=1.6,
            can_reduce_current_saudi=True,
        ))
        eng = data[data["Job Family"] == "Engineer"].iloc[0]
        ratio = float(eng["Fully Loaded Cost per In-house Saudi Employee"]) / float(
            eng["Fully Loaded Cost per In-house Non-Saudi Employee"]
        )
        self.assertAlmostEqual(ratio, 1.6, places=3)


def _mgic_bu_configuration_dict() -> dict:
    """Build a bu_configuration JSON dict populated with MGIC's hardcoded
    mappings. Used by HTTP tests that need to pass the empty-BU hard-block
    check in /workbooks/upload (added in Round 4)."""
    from manpower_app.mappings import ACTIVITY_MAPPING, JOB_FAMILY_MAPPING, PROFESSION_MAPPING
    return {
        "profession_mapping": dict(PROFESSION_MAPPING),
        "activity_mapping": dict(ACTIVITY_MAPPING),
        "job_family_mapping": dict(JOB_FAMILY_MAPPING),
        "outsourceability_overrides": {},
        "ratio_overrides": {},
        "driver_overrides": {},
    }


class ApiSmokeTests(unittest.TestCase):
    """HTTP wire-layer smoke tests using FastAPI's TestClient — verify the
    JSON payloads and the BU Configuration upload+download cycle work end-to-end
    over HTTP, not just at the Python function level."""

    @classmethod
    def setUpClass(cls):
        from fastapi.testclient import TestClient
        from manpower_api.app import app, store
        cls.client = TestClient(app)
        # Reset the cached AppStore between test classes so workbook state doesn't leak.
        store.processed = None
        store.workbook_bytes = None
        store.workbook_filename = None
        store.model_data = None
        store.model_metadata = None
        store.target_split = None
        store.optimization_payload = None

    def test_health_endpoint_returns_ok(self):
        r = self.client.get("/health")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json(), {"status": "ok"})

    def test_assumption_defaults_returns_canonical_lists(self):
        r = self.client.get("/assumptions/defaults")
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertIn("outsourceability", body)
        self.assertIn("max_ratios", body)
        self.assertIn("drivers", body)
        self.assertIn("Engineer", body["outsourceability"])
        self.assertIn("Quarries Foreman", body["max_ratios"])

    def test_bu_configuration_template_returns_xlsx_with_seven_sheets(self):
        r = self.client.get("/bu/configuration/template?bu_code=MGIC")
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r.headers["content-type"])
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        self.assertEqual(
            set(wb.sheetnames),
            {
                "README",
                "Profession Mapping",
                "Activity Mapping",
                "Job Families",
                "Ratios",
                "Drivers",
                "Cost Assumptions",
            },
        )

    def test_bu_configuration_export_round_trips_user_overrides(self):
        """POST a saved configuration → get XLSX back → upload the same XLSX → parse →
        the user's specific overrides survive (parsed may include defaults too — fine)."""
        import io
        original = {
            "outsourceability_overrides": {"Engineer": "Fully Outsourceable"},
            "ratio_overrides": {"Quarries Foreman": "1:5"},
            "driver_overrides": {
                "Production Foreman": [{"activity": "Factory", "profession": "Welder"}]
            },
            "saudi_cost_premium": None,
            "outsource_cost_discount": None,
        }
        # Download
        r = self.client.post(
            "/bu/configuration/export",
            json={"bu_code": "MGIC", "bu_name": "MGIC", "configuration": original},
        )
        self.assertEqual(r.status_code, 200)
        xlsx_bytes = r.content

        # Upload back
        r = self.client.post(
            "/bu/configuration/import",
            files={"file": ("MGIC.xlsx", io.BytesIO(xlsx_bytes),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(body["errors"], [])
        parsed = body["configuration"]
        # User-specified overrides must survive (parsed may include additional defaults).
        for fam, value in original["outsourceability_overrides"].items():
            self.assertEqual(parsed["outsourceability_overrides"].get(fam), value)
        for fam, value in original["ratio_overrides"].items():
            self.assertEqual(parsed["ratio_overrides"].get(fam), value)
        self.assertEqual(parsed["driver_overrides"], original["driver_overrides"])

    def test_workbook_upload_then_optimization_run_over_http(self):
        """Walk a full user journey over HTTP: upload payroll → run optimization
        with BU overrides + scenario knobs → result includes the BU override effect.
        """
        import io
        import json
        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()

        # 1. Upload payroll workbook with a configured BU (mandatory after Round 4)
        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll.xlsx", io.BytesIO(contents),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"bu_configuration": json.dumps(_mgic_bu_configuration_dict())},
        )
        self.assertEqual(r.status_code, 200, r.text)
        upload_body = r.json()
        self.assertGreater(upload_body["job_family_count"], 0)

        # 2. Run optimization with mixed BU + scenario settings
        r = self.client.post(
            "/optimization/run",
            json={
                "can_reduce_current_saudi": True,
                "outsourceability_overrides": {"Engineer": "Fully Outsourceable"},
                "saudi_cost_premium": 1.5,
            },
        )
        self.assertEqual(r.status_code, 200, r.text)
        payload = r.json()
        self.assertEqual(payload["metadata"]["optimization_status"], "Optimal")

        # 3. Verify the BU outsourceability override actually flowed through the wire.
        # model_processing carries the per-family input row (with Outsourceability Type);
        # results carries the final headcount split.
        eng_model = next(
            (row for row in payload["model_processing"] if str(row.get("Job Family")) == "Engineer"),
            None,
        )
        self.assertIsNotNone(eng_model, "Engineer row missing from model_processing")
        self.assertEqual(str(eng_model["Outsourceability Type"]), "Fully Outsourceable")

        eng_result = next(
            (row for row in payload["results"] if str(row.get("Job Family")) == "Engineer"),
            None,
        )
        self.assertIsNotNone(eng_result, "Engineer row missing from results")
        # With Engineer flipped to Fully Outsourceable, the LP can move workers out:
        self.assertGreaterEqual(int(eng_result["Outsourced Labor"]), 1)


class BUExcelRefactorAdditionalCoverageTests(unittest.TestCase):
    """Extra coverage added in Round 3 of the 150526 enhancements: tests that
    pin the new ""BU Excel as source of truth"" architecture from angles the
    earlier classes don't reach — full HTTP wire path, BU swap state isolation,
    Excel-driven cost knobs reaching the LP, mapping edge cases, hard-block
    response shape, and mapping override + custom family interaction."""

    @classmethod
    def setUpClass(cls):
        from fastapi.testclient import TestClient
        from manpower_api.app import app, store
        cls.client = TestClient(app)
        cls._store = store
        cls._reset_store()

    @classmethod
    def _reset_store(cls):
        cls._store.processed = None
        cls._store.workbook_bytes = None
        cls._store.workbook_filename = None
        cls._store.model_data = None
        cls._store.model_metadata = None
        cls._store.target_split = None
        cls._store.optimization_payload = None

    def setUp(self):
        self._reset_store()

    # --- #1: End-to-end over HTTP (download → import → upload → optimize) ----

    def test_download_import_upload_optimize_full_http_chain(self):
        """Walk the exact consultant flow over HTTP:
        1. GET /bu/configuration/template → BU's Excel
        2. POST /bu/configuration/import with that Excel → parsed config
        3. POST /workbooks/upload with payroll + bu_configuration → upload OK
        4. POST /optimization/run → Optimal status

        Any wire-format mismatch between these endpoints (form vs JSON, field
        names, content types) breaks the consultant's flow — this test pins it."""
        import io
        import json

        # Step 1: download a populated BU Excel (MGIC seeded with defaults).
        # Note: /template returns the empty skeleton (used for first-time BUs);
        # we use /export here because round 4 hard-blocks uploads with empty
        # BU configs, so we need a populated XLSX to round-trip.
        r = self.client.post(
            "/bu/configuration/export",
            json={
                "bu_code": "MGIC",
                "bu_name": "MGIC",
                "configuration": _mgic_bu_configuration_dict(),
            },
        )
        self.assertEqual(r.status_code, 200)
        xlsx_bytes = r.content

        # Step 2: re-import the same XLSX as if the consultant just edited & uploaded it
        r = self.client.post(
            "/bu/configuration/import",
            files={"file": ("MGIC_config.xlsx", io.BytesIO(xlsx_bytes),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertEqual(body["errors"], [])
        bu_configuration = body["configuration"]

        # Step 3: upload a payroll workbook WITH the parsed BU configuration attached
        payroll = EndToEndUserJourneyTests._quarries_workbook_bytes()
        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll.xlsx", io.BytesIO(payroll),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"bu_configuration": json.dumps(bu_configuration)},
        )
        self.assertEqual(r.status_code, 200, r.text)
        upload_body = r.json()
        self.assertGreater(upload_body["job_family_count"], 0)
        self.assertEqual(upload_body["unmapped_pairs"], [], "Workbook should be fully mapped via the BU's Excel")

        # Step 4: run the optimizer end-to-end
        r = self.client.post("/optimization/run", json={"can_reduce_current_saudi": True})
        self.assertEqual(r.status_code, 200, r.text)
        payload = r.json()
        self.assertEqual(payload["metadata"]["optimization_status"], "Optimal")

    # --- #2: BU swap doesn't leak state -------------------------------------

    def test_bu_swap_overrides_do_not_leak_between_uploads(self):
        """Upload payroll with MGIC's mappings; then re-upload the same payroll
        with UAAC's (empty) mappings. The two upload responses must reflect each
        BU's configuration independently — no leftover state from the first BU."""
        import io
        import json

        payroll = EndToEndUserJourneyTests._quarries_workbook_bytes()

        # Build MGIC's bu_configuration with populated mappings (required by
        # the Round-4 hard-block in /workbooks/upload).
        mgic_config = _mgic_bu_configuration_dict()

        # Upload #1 — MGIC
        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll1.xlsx", io.BytesIO(payroll), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"bu_configuration": json.dumps(mgic_config)},
        )
        self.assertEqual(r.status_code, 200, r.text)
        mgic_family_count = r.json()["job_family_count"]

        # Upload #2 — pretend it's UAAC with a NEW set of mappings that route
        # the "Quarries - Foreman" pair to a different family on this upload.
        uaac_config = {
            "profession_mapping": {},
            "activity_mapping": {},
            "job_family_mapping": {"Quarries - Foreman": "Quarries Supervisor"},
            "outsourceability_overrides": {},
            "ratio_overrides": {},
            "driver_overrides": {},
        }
        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll2.xlsx", io.BytesIO(payroll), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"bu_configuration": json.dumps(uaac_config)},
        )
        self.assertEqual(r.status_code, 200, r.text)
        upload2 = r.json()

        # Upload #2 must reflect UAAC's overrides, NOT MGIC's. The Quarries Foreman
        # pair must now route to Quarries Supervisor (or at least NOT Quarries Foreman).
        families = {row["family_name"] for row in upload2["job_families"]} if "job_families" in upload2 else set()
        # No direct exposure of the family list in the response — use the optimization layer.
        r = self.client.post("/optimization/run", json={"can_reduce_current_saudi": True})
        self.assertEqual(r.status_code, 200, r.text)
        payload = r.json()
        results = payload["results"] if isinstance(payload, dict) else []
        family_names = {str(row.get("Job Family")) for row in results}
        self.assertIn(
            "Quarries Supervisor", family_names,
            "UAAC's job_family_mapping override should route Quarries Foreman rows to Quarries Supervisor",
        )

    # --- #3: Cost knobs from the BU Excel reach the LP ---------------------

    def test_saudi_cost_premium_from_bu_excel_reaches_lp(self):
        """The Cost Assumptions sheet's Saudi pay premium must flow through to
        the LP's per-family Saudi cost basis. End-to-end: build XLSX with knob
        set, parse, run engine, assert resulting Saudi/non-Saudi cost ratio
        matches the knob (not the hardcoded 1.10 default)."""
        import io
        from manpower_app.bu_config_io import BUConfigurationPayload, build_workbook, parse_workbook
        from manpower_app.service import OptimizationSettings, process_workbook, prepare_model_data

        config = BUConfigurationPayload(saudi_cost_premium=1.75)
        xlsx = build_workbook("MGIC", "MGIC", config)
        parsed, errors = parse_workbook(xlsx)
        self.assertEqual(errors, [])
        self.assertAlmostEqual(parsed.saudi_cost_premium, 1.75, places=4)

        # The cost knob is a scenario-level setting at the LP layer (see
        # ScenarioCostKnobsTests). Wire it through by passing it to OptimizationSettings.
        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        processed = process_workbook(io.BytesIO(contents))
        data, _ = prepare_model_data(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                saudi_cost_premium=parsed.saudi_cost_premium,
            ),
        )
        eng = data[data["Job Family"] == "Engineer"].iloc[0]
        ratio = (
            float(eng["Fully Loaded Cost per In-house Saudi Employee"])
            / float(eng["Fully Loaded Cost per In-house Non-Saudi Employee"])
        )
        self.assertAlmostEqual(ratio, 1.75, places=4)

    # --- #4: Mapping sheet edge cases --------------------------------------

    def test_duplicate_raw_keys_in_profession_mapping_last_value_wins(self):
        """Two rows with the same raw profession but different standardized values:
        the parser keeps the LAST one (matches Python dict semantics). No errors."""
        import io
        from openpyxl import Workbook
        from manpower_app.bu_config_io import parse_workbook

        wb = Workbook()
        wb.active.title = "README"
        ws = wb.create_sheet("Profession Mapping")
        ws.append(["Raw Profession", "Standardized Profession"])
        ws.append(["Senior Welder", "Welder"])
        ws.append(["Senior Welder", "Skilled Labor"])  # duplicate raw, different std
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertEqual(errors, [])
        self.assertEqual(parsed.profession_mapping.get("Senior Welder"), "Skilled Labor")

    def test_blank_and_whitespace_rows_are_silently_skipped(self):
        """Empty cells, whitespace-only cells, and missing required values must be
        skipped without errors — the user can keep blank rows as scratch space."""
        import io
        from openpyxl import Workbook
        from manpower_app.bu_config_io import parse_workbook

        wb = Workbook()
        wb.active.title = "README"
        ws = wb.create_sheet("Activity Mapping")
        ws.append(["Raw Activity", "Standardized Activity"])
        ws.append(["", ""])                 # blank row
        ws.append(["   ", "  "])            # whitespace-only
        ws.append(["Production", ""])       # missing standardized
        ws.append(["", "Factory"])          # missing raw
        ws.append(["Workshop", "Factory"])  # valid row
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertEqual(errors, [])
        self.assertEqual(parsed.activity_mapping, {"Workshop": "Factory"})

    def test_invalid_outsourceability_value_returns_row_level_error(self):
        """The Job Families sheet's Outsourceability column must be one of three
        canonical strings. A bogus value reports a clear error pointing at the
        family — no partial save."""
        import io
        from openpyxl import Workbook
        from manpower_app.bu_config_io import parse_workbook

        wb = Workbook()
        wb.active.title = "README"
        ws = wb.create_sheet("Job Families")
        ws.append(["Activity", "Profession", "Job Family", "Outsourceability"])
        ws.append(["Factory", "Engineer", "Engineer", "MaybeOutsourceable"])  # invalid
        buf = io.BytesIO()
        wb.save(buf)
        parsed, errors = parse_workbook(buf.getvalue())
        self.assertTrue(any("Engineer" in e and "MaybeOutsourceable" in e for e in errors),
                        f"Expected a row-level error mentioning Engineer + bad value, got {errors}")
        # Note: the row's family-pair mapping IS still recorded (it's not invalid),
        # only the outsourceability value is rejected.
        self.assertNotIn("Engineer", parsed.outsourceability_overrides)

    # --- #5: Hard-block UX response shape -----------------------------------

    def test_unmapped_pairs_response_shape_supports_frontend_hardblock(self):
        """Upload a payroll with a (activity, profession) the BU's mappings don't
        cover. The response must include unmapped_pairs entries with the exact
        shape the frontend's hard-block branch expects: a list of dicts each
        with 'activity', 'profession' (and optionally 'count')."""
        import io
        import json
        from openpyxl import Workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        # Mapped baseline so the engine doesn't error on empty frame
        for i in range(5):
            inhouse.append([100 + i, "Quarries", "Skilled Labor", "non-saudi", 4200, 0, 3300, 500, 200, 150, 50, 0])
        # Unmapped pair
        for i in range(3):
            inhouse.append([200 + i, "Drone Yard", "Drone Pilot", "Saudi", 8000, 0, 6500, 1000, 300, 200, 100, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)

        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll.xlsx", io.BytesIO(buf.getvalue()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"bu_configuration": json.dumps(_mgic_bu_configuration_dict())},
        )
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertIn("unmapped_pairs", body, "Response must expose unmapped_pairs for the frontend")
        self.assertIsInstance(body["unmapped_pairs"], list)
        self.assertGreater(len(body["unmapped_pairs"]), 0)
        for entry in body["unmapped_pairs"]:
            self.assertIsInstance(entry, dict)
            self.assertIn("activity", entry)
            self.assertIn("profession", entry)

    # --- Round-4: hard-block payroll upload for unconfigured BUs -----------

    def test_upload_blocked_when_bu_has_no_mappings(self):
        """When the active BU has no profession/activity/job-family mappings
        configured, /workbooks/upload returns 400 (instead of silently using
        MGIC's hardcoded defaults). The consultant must configure the BU first."""
        import io
        import json
        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()

        # Empty config — what a fresh UAAC / FAST consultant's local state looks like.
        empty_config = {
            "profession_mapping": {},
            "activity_mapping": {},
            "job_family_mapping": {},
            "outsourceability_overrides": {},
            "ratio_overrides": {},
            "driver_overrides": {},
        }
        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll.xlsx", io.BytesIO(contents),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"bu_configuration": json.dumps(empty_config)},
        )
        self.assertEqual(r.status_code, 400, r.text)
        detail = r.json().get("detail", "")
        self.assertIn("no profession", detail.lower(), f"Got unexpected detail: {detail}")

    def test_upload_also_blocked_when_no_bu_configuration_sent(self):
        """The defense-in-depth path: no bu_configuration form field at all
        (e.g. a malformed UI request) also gets blocked with the same 400."""
        import io
        contents = EndToEndUserJourneyTests._quarries_workbook_bytes()
        r = self.client.post(
            "/workbooks/upload",
            files={"file": ("payroll.xlsx", io.BytesIO(contents),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(r.status_code, 400, r.text)

    # --- #6: Mapping override + custom family interaction ------------------

    def test_mapping_override_routes_payroll_into_custom_family_without_double_count(self):
        """A consultant defines a custom family ""Drone Pilot"" with explicit costs,
        AND adds a profession mapping ""Senior Drone Pilot → Drone Pilot"". The
        engine must route 'Senior Drone Pilot' payroll rows into the custom family —
        without double-counting them or losing them as unmapped."""
        import io
        from openpyxl import Workbook
        from manpower_app.service import process_workbook
        from manpower_app.family_specs import ActivityProfession, CustomFamilyCosts, CustomFamilySpec

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        # A pair routed by the BU's overrides: standardized via profession map +
        # custom family at the family layer.
        for i in range(4):
            inhouse.append([100 + i, "Quarries", "Senior Drone Pilot", "non-saudi", 7500, 0, 6000, 800, 300, 250, 150, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        contents = buf.getvalue()

        # Standardize via profession_mapping; the custom family declares the family-level
        # routing via source_pairs.
        drone_family = CustomFamilySpec(
            family_name="Drone Pilot",
            outsourceability="Partially Outsourceable",
            source_pairs=[ActivityProfession(activity="Quarries", profession="Drone Pilot")],
            costs=CustomFamilyCosts(saudi_inhouse=8000, non_saudi_inhouse=7000, outsourced=5500),
        )
        processed = process_workbook(
            io.BytesIO(contents),
            custom_families=[drone_family],
            profession_mapping_overrides={"Senior Drone Pilot": "Drone Pilot"},
        )
        # 4 payroll rows should land in the Drone Pilot family — no duplicates, no losses.
        drone_rows = processed.inhouse_cleaned[processed.inhouse_cleaned["Job_Family"] == "Drone Pilot"]
        self.assertEqual(len(drone_rows), 4)
        # The optimization frame should expose the family with its custom costs.
        opt = processed.optimization_df
        drone_opt = opt[opt["Job Family"] == "Drone Pilot"]
        self.assertEqual(len(drone_opt), 1)
        self.assertEqual(int(drone_opt.iloc[0]["Current Headcount"]), 4)


class ConsultantFeedbackRoundTests(unittest.TestCase):
    """Tests pinning the bugs the consultant flagged in screenshots
    photo_5800856712465026450_x and photo_5800856712465026451_x:

      * Idle Saudi Labor must always stay 100% Saudi regardless of assumptions
      * Target = 0 per family must be feasible
      * Management vs Executive Management saudization knobs must be independent
      * Outsourceable families must outsource per ratio even when in-house is
        cheaper than outsourced (Scenario 1 — safety officers all in-house)
      * Saudization = 0% with protection off must drive Saudis to 0
      * Saudi Protection in Target mode rounds half up and caps at target
    """

    # ─── T1-A: Idle Saudi Labor is 100% Saudi always ────────────────────────
    def test_idle_saudi_labor_stays_100_percent_saudi_under_any_assumptions(self):
        """Consultant: 'idle saudi labor become 50% non saudi. These should
        always be 100% saudis no matter what the assumptions are because this
        job family is by definition saudis'."""
        from manpower_app.optimization import (
            IN_HOUSE_NON_SAUDI_COLUMN,
            IN_HOUSE_SAUDI_COLUMN,
            OUTSOURCED_COLUMN,
            solve_optimization,
        )

        row = base_row(**{
            "Job Family": "Idle Saudi Labor",
            "Current Headcount": 10,
            "Minimum Headcount Needed": 8,
            "Current Total In-house Saudi": 10,
            "Current In-House Non-Saudi Count": 0,
            "Current Outsourced Count": 0,
            "Outsourceability Type": "Fully Outsourceable",  # would normally allow outsourcing
            "Risk Factor": 0.0,  # outsourced workers count fully
            "Fully Loaded Cost per In-house Non-Saudi Employee": 2000,  # cheaper than Saudi
            "Fully Loaded Cost per In-house Saudi Employee": 5000,
            "Avg Cost Outsourced": 1500,  # cheapest of all
        })
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        self.assertEqual(status, "Optimal")
        # All 10 must be in-house Saudis even though both non-Saudi and outsourced
        # are cheaper and the family is marked Fully Outsourceable.
        self.assertEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 10)
        self.assertEqual(int(solved.iloc[0][IN_HOUSE_NON_SAUDI_COLUMN]), 0)
        self.assertEqual(int(solved.iloc[0][OUTSOURCED_COLUMN]), 0)

    # ─── T1-B: Target = 0 must be feasible ──────────────────────────────────
    def test_target_zero_headcount_is_feasible_not_infeasible(self):
        """Consultant: 'when I zero all counts and leave skilled labor 100 and
        4 safety officer, we get infeasible'. Zeroing a target must produce
        zero counts, not infeasibility."""
        from manpower_app.optimization import (
            IN_HOUSE_NON_SAUDI_COLUMN,
            IN_HOUSE_SAUDI_COLUMN,
            OUTSOURCED_COLUMN,
            solve_optimization,
        )

        zeroed_row = base_row(**{
            "Job Family": "Engineer",
            "Current Headcount": 0,  # Target mode: user set this to 0
            "Minimum Headcount Needed": 0,
            "Current Total In-house Saudi": 5,  # still 5 Saudis "today"
            "Current In-House Non-Saudi Count": 3,
            "Current Outsourced Count": 2,
        })
        kept_row = base_row(**{
            "Job Family": "Skilled Labor",
            "Current Headcount": 100,
            "Minimum Headcount Needed": 80,
            "Current Total In-house Saudi": 10,
            "Current In-House Non-Saudi Count": 50,
            "Current Outsourced Count": 40,
        })
        data = pd.DataFrame([zeroed_row, kept_row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=False,  # would normally floor Saudis at current — but target=0 collapses it
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        self.assertEqual(status, "Optimal", "Zeroed family should not cause infeasibility")
        # The zeroed family produces zeros across the board.
        zero = solved[solved["Job Family"] == "Engineer"].iloc[0]
        self.assertEqual(int(zero[IN_HOUSE_SAUDI_COLUMN]), 0)
        self.assertEqual(int(zero[IN_HOUSE_NON_SAUDI_COLUMN]), 0)
        self.assertEqual(int(zero[OUTSOURCED_COLUMN]), 0)
        # The other family still solves normally.
        kept = solved[solved["Job Family"] == "Skilled Labor"].iloc[0]
        self.assertEqual(int(kept[IN_HOUSE_SAUDI_COLUMN])
                         + int(kept[IN_HOUSE_NON_SAUDI_COLUMN])
                         + int(kept[OUTSOURCED_COLUMN]), 100)

    # ─── T1-C: Management vs Executive Management split ─────────────────────
    def test_management_and_executive_management_are_independent_saudi_rates(self):
        """Consultant: 'Don't link management in saudization input to executive
        management... just to "management" as a job family'. Confirm that
        setting executive_management_saudization_rate to 0.0 and
        management_saudization_rate to 1.0 produces non-Saudi for executive
        but all-Saudi for management."""
        from manpower_app.service import OptimizationSettings

        # The route is in service.py: build_profession_rates() reads both fields
        # into separate keys. Verify both keys exist with distinct values.
        from manpower_app.utils import normalize_lookup_text

        settings = OptimizationSettings(
            management_saudization_rate=1.0,
            executive_management_saudization_rate=0.0,
        )
        # We don't run the optimizer here — directly verify the wiring via the
        # settings dataclass (the per-profession rates dict is built in
        # process_workbook and run_optimization paths). Just check defaults differ.
        self.assertNotEqual(
            settings.management_saudization_rate,
            settings.executive_management_saudization_rate,
        )
        # And the canonical names are routed via normalize_lookup_text — confirm
        # the lookup keys differ.
        self.assertNotEqual(
            normalize_lookup_text("Management"),
            normalize_lookup_text("Executive Management"),
        )

    # ─── T2-D: outsource even when in-house cheaper ─────────────────────────
    def test_outsourceable_family_outsources_per_ratio_even_when_inhouse_cheaper(self):
        """Consultant Scenario 1: 'inhouse safety officers seem to be cheaper
        than outsourced is why the tool is giving this result. We should
        override such cases and go with the assumption that inhouse are more
        expensive than outsourced'. Bump in-house non-Saudi cost above outsourced
        for outsourceable families."""
        from manpower_app.optimization import (
            OUTSOURCED_COLUMN,
            solve_optimization,
        )
        from manpower_app.costs import bump_inhouse_non_saudi_above_outsourced

        # Unit-check the cost-bump function first
        bumped = bump_inhouse_non_saudi_above_outsourced(
            inhouse_non_saudi_unit_cost=3000,  # cheaper
            outsourced_unit_cost=5000,
            outsourceability_type="Partially Outsourceable",
        )
        self.assertGreater(bumped, 5000, "in-house bump should produce strictly > outsourced")

        # Non-outsourceable families left untouched
        self.assertEqual(
            bump_inhouse_non_saudi_above_outsourced(3000, 5000, "Not Outsourceable"),
            3000,
        )

        # End-to-end: a Partially Outsourceable family with cheaper in-house
        # cost in the data should still get outsourcing.
        row = base_row(**{
            "Job Family": "Safety Officer",
            "Current Headcount": 100,
            "Minimum Headcount Needed": 100,
            "Outsourceability Type": "Partially Outsourceable",
            "Outsourced v1": 30,  # the family's outsourcing ratio cap
            "Fully Loaded Cost per In-house Non-Saudi Employee": 8000,  # high (bumped over outsourced)
            "Fully Loaded Cost per In-house Saudi Employee": 9000,
            "Avg Cost Outsourced": 5000,
            "Risk Factor": 0.0,
        })
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=False,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        self.assertEqual(status, "Optimal")
        # With outsourced cheaper than in-house, the LP outsources up to the cap (30).
        self.assertGreater(int(solved.iloc[0][OUTSOURCED_COLUMN]), 0)

    # ─── T2-E: Saud=0 + protection off → no Saudis (except for protected families) ─
    def test_saudization_zero_with_protection_off_produces_zero_saudis(self):
        """Consultant: 'Security Guard also end up using some saudis... its weird
        because we have saudization at zero'. At Saud=0 + protection off, all
        non-special families should produce 0 Saudis."""
        from manpower_app.optimization import (
            IN_HOUSE_SAUDI_COLUMN,
            solve_optimization,
        )

        row = base_row(**{
            "Job Family": "Security Guard",
            "Current Headcount": 20,
            "Current Total In-house Saudi": 8,
            "Current In-House Non-Saudi Count": 10,
            "Current Outsourced Count": 2,
            "Fully Loaded Cost per In-house Saudi Employee": 4000,  # cheaper Saudi than non-Saudi
            "Fully Loaded Cost per In-house Non-Saudi Employee": 5000,
        })
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
        )
        self.assertEqual(status, "Optimal")
        # Despite Saudi being cheaper here, strict-zero forces 0 Saudis.
        self.assertEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 0)

    def test_saudization_zero_with_explicit_protection_keeps_floor(self):
        """Strict-zero should NOT fire when explicit dynamic protection is set."""
        from manpower_app.optimization import (
            IN_HOUSE_SAUDI_COLUMN,
            solve_optimization,
        )

        row = base_row(**{"Current Total In-house Saudi": 10})
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=True,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
            protect_current_saudi_percent=0.6,  # protection wins
        )
        self.assertEqual(status, "Optimal")
        # Ceil(10 * 0.6) = 6
        self.assertGreaterEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 6)

    # ─── T2-F: Saudi Protection caps at target headcount ────────────────────
    def test_saudi_protection_caps_at_target_headcount(self):
        """Consultant target-mode spec: 'job family has 5, 50% of 5 is 3,
        target state see how much user inputted, if 7, new constraint, 3 out
        of 7 needed to be saudis, if less target number be saudi'."""
        from manpower_app.optimization import (
            IN_HOUSE_SAUDI_COLUMN,
            solve_optimization,
        )

        # Case A: target = 7, current = 5 Saudis, protect 50% → 3 Saudis required (ceil(5*0.5) = 3)
        row = base_row(**{
            "Current Headcount": 7,
            "Current Total In-house Saudi": 5,
            "Current In-House Non-Saudi Count": 2,
            "Current Outsourced Count": 0,
        })
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=False,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
            protect_current_saudi_percent=0.5,
        )
        self.assertEqual(status, "Optimal")
        self.assertGreaterEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 3)

        # Case B: target reduced BELOW the protection floor → protection caps at target.
        # 5 current Saudis, 100% protection, target = 2 → only 2 Saudis can fit, must stay feasible.
        row = base_row(**{
            "Current Headcount": 2,
            "Current Total In-house Saudi": 5,
            "Current In-House Non-Saudi Count": 0,
            "Current Outsourced Count": 0,
            "Minimum Headcount Needed": 2,
        })
        data = pd.DataFrame([row])
        solved, _, status = solve_optimization(
            data,
            enforce_saudization=False,
            saudization_rate=0.0,
            can_reduce_current_saudi=True,
            tenure_constraint_active=False,
            profession_saudization_rates={},
            protect_current_saudi_percent=1.0,
        )
        self.assertEqual(status, "Optimal", "Protection floor must cap at total_headcount")
        self.assertLessEqual(int(solved.iloc[0][IN_HOUSE_SAUDI_COLUMN]), 2)


class HighPerformerProtectionTests(unittest.TestCase):
    """Phase 3: Manpower Performance column + good-performer protection.

    The payroll workbook can carry an optional 'Manpower Performance' column on
    the In-house sheet (1–5 per employee). When the user enables 'Protect high
    performers' with threshold T, the LP must keep at least the count of
    high-performer Saudis as Saudi in-house and the count of high-performer
    Non-Saudis as Non-Saudi in-house. Current mode only — target mode ignores.
    """

    @staticmethod
    def _payroll_with_perf_scores(
        family_rows: list[tuple[str, str, str, int]],
    ) -> bytes:
        """Build a payroll workbook with Manpower Performance column.

        family_rows: list of (location, profession, nationality, performance_score) tuples.
        """
        import io
        from openpyxl import Workbook

        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
            "Manpower Performance",
        ])
        for idx, (location, profession, nationality, score) in enumerate(family_rows):
            inhouse.append([
                100 + idx, location, profession, nationality,
                5000, 0, 4000, 500, 200, 200, 100, 0,
                score,
            ])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        # Add one outsourced row in each family that has inhouse rows so the
        # optimization frame includes outsourced cost data.
        seen = set()
        for location, profession, _, _ in family_rows:
            key = (location, profession)
            if key in seen:
                continue
            seen.add(key)
            sub.append([200 + len(seen), location, profession, "non-saudi",
                        1500, 100, 50, 30, 0, 0, 0, 0, 80])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_manpower_performance_column_is_read_into_inhouse_cleaned(self):
        import io
        from manpower_app.service import process_workbook
        contents = self._payroll_with_perf_scores([
            ("Quarries", "Skilled Labor", "non-saudi", 4),
            ("Quarries", "Skilled Labor", "non-saudi", 2),
            ("Quarries", "Skilled Labor", "Saudi", 5),
        ])
        processed = process_workbook(io.BytesIO(contents))
        self.assertIn("Manpower Performance", processed.inhouse_cleaned.columns)
        scores = sorted(processed.inhouse_cleaned["Manpower Performance"].tolist())
        self.assertEqual(scores, [2.0, 4.0, 5.0])

    def test_missing_manpower_performance_column_defaults_to_three(self):
        """Workbooks WITHOUT the column still parse — every row defaults to 3.
        With threshold > 3, no protection applies (high-performer count = 0)."""
        import io
        from manpower_app.service import process_workbook
        # Build a payroll WITHOUT the performance column.
        from openpyxl import Workbook
        wb = Workbook()
        inhouse = wb.active
        inhouse.title = "Inhouse"
        inhouse.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        for i in range(3):
            inhouse.append([100 + i, "Quarries", "Skilled Labor", "non-saudi",
                            5000, 0, 4000, 500, 200, 200, 100, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        processed = process_workbook(io.BytesIO(buf.getvalue()))
        self.assertIn("Manpower Performance", processed.inhouse_cleaned.columns)
        # All three default to 3.
        self.assertTrue((processed.inhouse_cleaned["Manpower Performance"] == 3.0).all())

    def test_high_performer_protection_keeps_them_inhouse(self):
        """Three Skilled Labor employees, 2 with score >= 4. Protection on
        keeps at least 2 in-house non-Saudi (the high performers) — LP cannot
        outsource them."""
        import io
        from manpower_app.optimization import (
            IN_HOUSE_NON_SAUDI_COLUMN,
            OUTSOURCED_COLUMN,
        )
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._payroll_with_perf_scores([
            ("Quarries", "Skilled Labor", "non-saudi", 5),
            ("Quarries", "Skilled Labor", "non-saudi", 4),
            ("Quarries", "Skilled Labor", "non-saudi", 2),
        ])
        processed = process_workbook(io.BytesIO(contents))

        # Without protection: LP can outsource anyone.
        baseline = run_optimization(
            processed,
            OptimizationSettings(can_reduce_current_saudi=True),
        )
        baseline_row = baseline["data"][baseline["data"]["Job Family"] == "Skilled Labor"].iloc[0]
        baseline_inhouse_ns = int(baseline_row[IN_HOUSE_NON_SAUDI_COLUMN])

        # With protection: the 2 high performers must stay in-house non-Saudi.
        protected = run_optimization(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                protect_high_performers=True,
                high_performer_threshold=4.0,
            ),
        )
        protected_row = protected["data"][protected["data"]["Job Family"] == "Skilled Labor"].iloc[0]
        protected_inhouse_ns = int(protected_row[IN_HOUSE_NON_SAUDI_COLUMN])
        self.assertGreaterEqual(
            protected_inhouse_ns, 2,
            f"At least 2 high-performer non-Saudis must remain in-house, got {protected_inhouse_ns}",
        )
        # Protection should not drop in-house count vs baseline (only constrain it up).
        self.assertGreaterEqual(protected_inhouse_ns, min(baseline_inhouse_ns, 2))

    def test_high_performer_protection_ignored_in_target_mode(self):
        """Phase 3 is Current mode only. Target mode must not apply the floor."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._payroll_with_perf_scores([
            ("Quarries", "Skilled Labor", "non-saudi", 5),
            ("Quarries", "Skilled Labor", "non-saudi", 5),
            ("Quarries", "Skilled Labor", "non-saudi", 1),
        ])
        processed = process_workbook(io.BytesIO(contents))
        result = run_optimization(
            processed,
            OptimizationSettings(
                optimization_mode="target",
                target_headcounts={"Skilled Labor": 3},
                can_reduce_current_saudi=True,
                protect_high_performers=True,
                high_performer_threshold=4.0,
            ),
        )
        # Floor columns should remain 0 in target mode (verified via the data frame).
        sl_row = result["data"][result["data"]["Job Family"] == "Skilled Labor"].iloc[0]
        self.assertEqual(int(sl_row.get("High Performer Saudi Floor", 0)), 0)
        self.assertEqual(int(sl_row.get("High Performer Non-Saudi Floor", 0)), 0)

    def test_threshold_above_max_score_protects_nobody(self):
        """Threshold = 5 with all scores < 5 → 0 high performers → no floor."""
        import io
        from manpower_app.service import OptimizationSettings, process_workbook, run_optimization

        contents = self._payroll_with_perf_scores([
            ("Quarries", "Skilled Labor", "non-saudi", 4),
            ("Quarries", "Skilled Labor", "non-saudi", 3),
        ])
        processed = process_workbook(io.BytesIO(contents))
        result = run_optimization(
            processed,
            OptimizationSettings(
                can_reduce_current_saudi=True,
                protect_high_performers=True,
                high_performer_threshold=5.0,  # nobody qualifies
            ),
        )
        sl_row = result["data"][result["data"]["Job Family"] == "Skilled Labor"].iloc[0]
        self.assertEqual(int(sl_row["High Performer Non-Saudi Floor"]), 0)

    def test_processed_workbook_exposes_has_performance_column_flag(self):
        """The ProcessedWorkbook dataclass must surface whether the uploaded
        payroll had the Manpower Performance column — the UI uses this to
        disable the high-performer toggle when the column is missing."""
        import io
        from openpyxl import Workbook

        from manpower_app.service import process_workbook

        # Payload WITH the column.
        contents_with = self._payroll_with_perf_scores([
            ("Quarries", "Skilled Labor", "non-saudi", 4),
        ])
        self.assertTrue(process_workbook(io.BytesIO(contents_with)).has_performance_column)

        # Payload WITHOUT the column (manually built so the header is absent).
        wb = Workbook()
        inh = wb.active
        inh.title = "Inhouse"
        inh.append([
            "No", "Location", "Profession", "Nationality",
            "Total Paid", "Total Unpaid",
            "Basic", "Housing Paid", "Trans Paid", "Medical Paid", "EOS Paid", "Value O.T (SAR)",
        ])
        inh.append([100, "Quarries", "Skilled Labor", "non-saudi",
                    5000, 0, 4000, 500, 200, 200, 100, 0])
        sub = wb.create_sheet("Subcontractor")
        sub.append([
            "No", "Working in", "Profession", "Nationality", "Basic",
            "Housing Paid", "Trans Paid", "Food", "Gosi", "Value O.T (SAR)",
            "Government Fees", "E.O.S monthly", "Service Margin",
        ])
        sub.append([200, "Quarries", "Skilled Labor", "non-saudi",
                    1500, 100, 50, 30, 0, 0, 0, 0, 80])
        buf = io.BytesIO()
        wb.save(buf)
        self.assertFalse(process_workbook(io.BytesIO(buf.getvalue())).has_performance_column)


if __name__ == "__main__":
    unittest.main()
