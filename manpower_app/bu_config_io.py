"""
Excel template export/import for per-BU Configuration.

The BU's Excel is the single source of truth for everything BU-specific:
the mapping pipeline (Profession → Activity → Job Family), the outsourceability
classification per family, supervisor:worker ratios, ratio drivers, and the
cost-assumption knobs (Saudi pay premium, outsource cost discount).

For a fresh BU the workbook is a "starter": every mapping row is pre-filled
with the tool's hardcoded defaults from ``manpower_app/mappings.py`` and
``manpower_app/rules.py`` so the consultant has a complete editing surface.
The user edits whichever rows differ for their BU and re-uploads.

For an already-configured BU the workbook is a snapshot of the saved values.

Round trip:
    config_json -> build_workbook(...) -> .xlsx bytes -> parse_workbook(...) -> config_json

Backwards-compatible: PR-#2-era 4-sheet workbooks (Outsourceability + Ratios +
Drivers + optional Engine Knobs) still parse. Missing new sheets simply fall
back to the engine's hardcoded defaults.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from manpower_app.mappings import (
    ACTIVITY_MAPPING,
    JOB_FAMILY_MAPPING,
    PROFESSION_MAPPING,
)
from manpower_app.rules import MAXIMUM_RATIO_RULES, OUTSOURCEABILITY_RULES


VALID_OUTSOURCEABILITY = {"Fully Outsourceable", "Partially Outsourceable", "Not Outsourceable"}
RATIO_PATTERN = re.compile(r"^\s*1\s*:\s*\d+\s*$")

DRIVER_SUPERVISORS = sorted([
    "Quarries Foreman",
    "Production Foreman",
    "Installation Foreman",
    "Showroom Supervisor",
    "Safety Officer",
    "Quarries Supervisor",
    "Installation Supervisor",
    "Factory Supervisor",
])

COST_ASSUMPTIONS = [
    ("Saudi pay premium", "How much more Saudi employees cost than non-Saudis (e.g. 1.10 = 10% more). Must be between 1.0 and 3.0. Blank = use the tool's default (1.10)."),
    ("Outsource cost discount", "Fraction outsourced labor is cheaper than non Saudi in-house (between 0.0 and 1.0). Blank = use the workbook's per-family value."),
]


@dataclass
class BUConfigurationPayload:
    """Snapshot of what a BU's Excel carries. Matches the desktop BUConfiguration type."""
    # Mapping pipeline (new — were hardcoded in mappings.py before this MR)
    profession_mapping: dict[str, str] = field(default_factory=dict)  # raw → standardized
    activity_mapping: dict[str, str] = field(default_factory=dict)    # raw → standardized
    job_family_mapping: dict[str, str] = field(default_factory=dict)  # "Activity - Profession" → family

    # Per-family / per-supervisor constraints
    outsourceability_overrides: dict[str, str] = field(default_factory=dict)
    ratio_overrides: dict[str, str] = field(default_factory=dict)
    driver_overrides: dict[str, list[dict[str, str]]] = field(default_factory=dict)

    # Cost assumptions (BU-level baselines)
    saudi_cost_premium: float | None = None
    outsource_cost_discount: float | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "profession_mapping": dict(self.profession_mapping),
            "activity_mapping": dict(self.activity_mapping),
            "job_family_mapping": dict(self.job_family_mapping),
            "outsourceability_overrides": dict(self.outsourceability_overrides),
            "ratio_overrides": dict(self.ratio_overrides),
            "driver_overrides": {k: list(v) for k, v in self.driver_overrides.items()},
            "saudi_cost_premium": self.saudi_cost_premium,
            "outsource_cost_discount": self.outsource_cost_discount,
        }

    @classmethod
    def from_dict(cls, payload: dict[str, Any] | None) -> "BUConfigurationPayload":
        payload = payload or {}
        return cls(
            profession_mapping=dict(payload.get("profession_mapping") or {}),
            activity_mapping=dict(payload.get("activity_mapping") or {}),
            job_family_mapping=dict(payload.get("job_family_mapping") or {}),
            outsourceability_overrides=dict(payload.get("outsourceability_overrides") or {}),
            ratio_overrides=dict(payload.get("ratio_overrides") or {}),
            driver_overrides={
                k: list(v) for k, v in (payload.get("driver_overrides") or {}).items()
            },
            saudi_cost_premium=payload.get("saudi_cost_premium"),
            outsource_cost_discount=payload.get("outsource_cost_discount"),
        )

    def is_empty(self) -> bool:
        return (
            not self.profession_mapping
            and not self.activity_mapping
            and not self.job_family_mapping
            and not self.outsourceability_overrides
            and not self.ratio_overrides
            and not self.driver_overrides
            and self.saudi_cost_premium is None
            and self.outsource_cost_discount is None
        )


def build_workbook(
    bu_code: str,
    bu_name: str | None,
    config: BUConfigurationPayload,
) -> bytes:
    """Return XLSX bytes — 7 sheets covering the full BU configuration surface.

    For an empty BU the mapping sheets are pre-filled with the tool's hardcoded
    defaults so consultants always have a complete reference to edit. For a
    populated BU the sheets show the saved values verbatim."""
    wb = openpyxl.Workbook()
    _build_readme_sheet(wb.active, bu_code, bu_name, config.is_empty())
    _build_profession_mapping_sheet(wb.create_sheet("Profession Mapping"), config)
    _build_activity_mapping_sheet(wb.create_sheet("Activity Mapping"), config)
    _build_job_families_sheet(wb.create_sheet("Job Families"), config)
    _build_ratios_sheet(wb.create_sheet("Ratios"), config)
    _build_drivers_sheet(wb.create_sheet("Drivers"), config)
    _build_cost_assumptions_sheet(wb.create_sheet("Cost Assumptions"), config)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_workbook(source: bytes | io.BytesIO) -> tuple[BUConfigurationPayload, list[str]]:
    """Return (config, errors). When ``errors`` is non-empty the caller should NOT save.

    Reads every recognized sheet. Older 4-sheet workbooks (Outsourceability + Ratios +
    Drivers + optional Engine Knobs) still parse — missing new sheets fall back to
    hardcoded defaults at the engine level."""
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    try:
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as exc:
        return BUConfigurationPayload(), [f"Could not open workbook: {exc}"]

    errors: list[str] = []
    config = BUConfigurationPayload()

    # Optional new sheets (mapping pipeline) — back-compat with PR-#2 workbooks.
    if "Profession Mapping" in wb.sheetnames:
        _parse_profession_mapping(wb["Profession Mapping"], config, errors)
    if "Activity Mapping" in wb.sheetnames:
        _parse_activity_mapping(wb["Activity Mapping"], config, errors)
    if "Job Families" in wb.sheetnames:
        _parse_job_families(wb["Job Families"], config, errors)

    # Required by PR-#2 workbooks; still expected for newer ones too.
    if "Outsourceability" in wb.sheetnames:
        # Legacy standalone sheet — pre-Job-Families merge. Still honored.
        _parse_outsourceability(wb["Outsourceability"], config, errors)
    if "Ratios" in wb.sheetnames:
        _parse_ratios(wb["Ratios"], config, errors)
    if "Drivers" in wb.sheetnames:
        _parse_drivers(wb["Drivers"], config, errors)

    # Cost knobs: new "Cost Assumptions" sheet OR legacy "Engine Knobs" sheet.
    if "Cost Assumptions" in wb.sheetnames:
        _parse_cost_assumptions(wb["Cost Assumptions"], config, errors)
    elif "Engine Knobs" in wb.sheetnames:
        _parse_cost_assumptions(wb["Engine Knobs"], config, errors)

    return config, errors


# ──────────────────────────────────────────────────────────────────────────
# Helpers — building each sheet


_HEADER_FILL = PatternFill(start_color="336B24", end_color="336B24", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_INSTRUCTION_FONT = Font(italic=True, color="475264")


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")


_EXAMPLE_FONT = Font(italic=True, color="8a96a3")


def _append_example_row(ws, *values: Any):
    """Append a row formatted as an editable example (italic + muted grey).
    Used in skeleton sheets so the consultant has a format hint they can
    overwrite with their own data."""
    ws.append(list(values))
    last_row = ws.max_row
    for col_idx in range(1, len(values) + 1):
        ws.cell(row=last_row, column=col_idx).font = _EXAMPLE_FONT


def _build_readme_sheet(ws, bu_code: str, bu_name: str | None, is_empty: bool):
    ws.title = "README"
    ws["A1"] = "Business Unit"
    ws["B1"] = bu_code
    ws["A2"] = "Display name"
    ws["B2"] = bu_name or ""
    if is_empty:
        instructions = (
            f"Skeleton workbook for {bu_code}. The mapping sheets are EMPTY (with a couple of "
            f"italic example rows for guidance). Fill them in with the values for THIS BU.\n\n"
            "Workflow (matches the consultant's original approach):\n\n"
            "1) Profession Mapping. List every unique profession that appears in your "
            "payroll and map it to a standardized profession name.\n"
            "2) Activity Mapping. Same, for the location / area values in your payroll.\n"
            "3) Job Families. For each unique (standardized Activity, standardized "
            "Profession) pair, pick which canonical job family it belongs to and set "
            "the outsourceability.\n"
            "4) Ratios. Set the max supervisor:worker ratio for each supervisor family.\n"
            "5) Drivers. List which (Activity, Profession) pairs count toward each "
            "supervisor.\n"
            "6) Cost Assumptions. Saudi pay premium + outsource cost discount.\n\n"
            "Anything left blank falls back to the tool's defaults at engine time. The "
            "Job Families sheet includes a side reference of the canonical family names "
            "and their default outsourceability so you know which families to route your "
            "pairs into."
        )
    else:
        instructions = (
            "Current saved configuration for this BU. Edit any value to change it; clear "
            "a value to fall back to the tool's default. Save and upload to apply."
        )
    ws["A4"] = instructions
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A4"].font = _INSTRUCTION_FONT
    ws.merge_cells("A4:E18")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40


def _build_profession_mapping_sheet(ws, config: BUConfigurationPayload):
    ws.append(["Profession (in payroll)", "Standardized Profession"])
    _style_header_row(ws)

    if config.profession_mapping:
        # Filled config: show user data merged with hardcoded defaults so the
        # consultant has full visibility.
        rows = {**PROFESSION_MAPPING, **config.profession_mapping}
        for payroll_value, standardized in sorted(rows.items()):
            ws.append([payroll_value, standardized])
    else:
        # Empty config (e.g. UAAC, FAST). Emit a skeleton: 2 example rows in italic
        # so the user has a format hint, then plenty of blank rows to fill.
        _append_example_row(ws, "(example) Junior Welder", "Welder")
        _append_example_row(ws, "(example) Senior Foreman", "Foreman")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    ws["D1"] = (
        "Fill this with your BU's data:\n"
        "1) Extract the unique profession values from your payroll.\n"
        "2) For each, choose a STANDARDIZED profession name.\n"
        "3) Replace the (example) rows and add one row per value.\n"
        "Blank entries fall back to the tool's defaults."
    )
    ws["D1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["D1"].font = _INSTRUCTION_FONT
    ws.merge_cells("D1:E8")
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 20


def _build_activity_mapping_sheet(ws, config: BUConfigurationPayload):
    ws.append(["Activity (in payroll)", "Standardized Activity"])
    _style_header_row(ws)

    if config.activity_mapping:
        rows = {**ACTIVITY_MAPPING, **config.activity_mapping}
        for payroll_value, standardized in sorted(rows.items()):
            ws.append([payroll_value, standardized])
    else:
        _append_example_row(ws, "(example) Production Plant", "Factory")
        _append_example_row(ws, "(example) Warehouse North", "Installation")

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 28
    ws["D1"] = (
        "Fill this with your BU's data:\n"
        "1) Extract the unique location / area values from your payroll.\n"
        "2) For each, choose a STANDARDIZED activity name.\n"
        "3) Replace the (example) rows and add one row per value.\n"
        "Blank entries fall back to the tool's defaults."
    )
    ws["D1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["D1"].font = _INSTRUCTION_FONT
    ws.merge_cells("D1:E8")
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 20


def _build_job_families_sheet(ws, config: BUConfigurationPayload):
    """Combined mapping + outsourceability sheet.

    Columns: Activity (standardized) | Profession (standardized) | Job Family |
             Outsourceability (one of the 3 valid strings).
    """
    ws.append(["Activity", "Profession", "Job Family", "Outsourceability"])
    _style_header_row(ws)

    # Emit data rows when there's anything to preserve: a job_family_mapping OR
    # outsourceability_overrides. Otherwise it's a skeleton (empty BU).
    has_data = bool(config.job_family_mapping or config.outsourceability_overrides)

    if has_data:
        rows = {**JOB_FAMILY_MAPPING, **config.job_family_mapping}
        for key, family in sorted(rows.items()):
            parts = re.split(r"\s*-\s*", key, maxsplit=1)
            if len(parts) == 2:
                activity, profession = parts[0].strip(), parts[1].strip()
            else:
                activity, profession = key.strip(), ""
            outsourceability = (
                config.outsourceability_overrides.get(family)
                or OUTSOURCEABILITY_RULES.get(family, "")
            )
            ws.append([activity, profession, family, outsourceability])
    else:
        # Skeleton: a couple of example rows showing the format.
        _append_example_row(
            ws,
            "(example) Factory", "(example) Welder", "Skilled Labor", "Fully Outsourceable",
        )
        _append_example_row(
            ws,
            "(example) Head Office", "(example) Accountant", "Administration", "Not Outsourceable",
        )

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 26

    # Right-side reference: canonical Job Family names + default outsourceability.
    # Consultants need this so they know which families to route their pairs into.
    ws["F1"] = "Canonical Job Families (reference)"
    ws["F1"].font = Font(bold=True, color="475264")
    ws["G1"] = "Default Outsourceability"
    ws["G1"].font = Font(bold=True, color="475264")
    for i, family_name in enumerate(sorted(OUTSOURCEABILITY_RULES.keys()), start=2):
        ws.cell(row=i, column=6, value=family_name).font = _INSTRUCTION_FONT
        ws.cell(row=i, column=7, value=OUTSOURCEABILITY_RULES[family_name]).font = _INSTRUCTION_FONT
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 26

    # Below the family reference: valid outsourceability values.
    valid_start = len(OUTSOURCEABILITY_RULES) + 3
    ws.cell(row=valid_start, column=6, value="Valid outsourceability values:").font = Font(
        bold=True, color="475264"
    )
    for i, value in enumerate(sorted(VALID_OUTSOURCEABILITY), start=valid_start + 1):
        ws.cell(row=i, column=6, value=value).font = _INSTRUCTION_FONT


def _build_ratios_sheet(ws, config: BUConfigurationPayload):
    ws.append(["Supervisor Family", "Value (e.g. 1:10)"])
    _style_header_row(ws)

    # Ratios sheet always lists the canonical supervisor families (their names are
    # universal — they're the same engine-recognized supervisor types for every BU).
    # For a skeleton: blank Value cells so the user fills their own. For a filled
    # config: show the saved values.
    for family in sorted(MAXIMUM_RATIO_RULES.keys()):
        saved = config.ratio_overrides.get(family)
        ws.append([family, saved if saved else ""])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws["D1"] = (
        "Set the maximum supervisor:worker ratio for each supervisor family. "
        "Format: 1:N (e.g. 1:10 means one supervisor for every 10 workers). "
        "Blank means the tool's default applies."
    )
    ws["D1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["D1"].font = _INSTRUCTION_FONT
    ws.merge_cells("D1:E5")
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 20


def _build_drivers_sheet(ws, config: BUConfigurationPayload):
    ws.append(["Supervisor Family", "Activity", "Profession"])
    _style_header_row(ws)
    if config.driver_overrides:
        for family in sorted(config.driver_overrides.keys()):
            for entry in config.driver_overrides[family]:
                ws.append([
                    family,
                    entry.get("activity", ""),
                    entry.get("profession", ""),
                ])
    else:
        # Starter: list supervisor names with blank Activity/Profession rows; user adds
        # one row per (activity, profession) pair counted toward that supervisor.
        for family in DRIVER_SUPERVISORS:
            ws.append([family, "", ""])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws["E1"] = "Add one row per Activity + Profession pair you want counted toward each supervisor's driver value"
    ws["E1"].font = _INSTRUCTION_FONT
    ws.column_dimensions["E"].width = 80


def _build_cost_assumptions_sheet(ws, config: BUConfigurationPayload):
    ws.append(["Knob", "Value", "Notes"])
    _style_header_row(ws)
    value_lookup = {
        "Saudi pay premium": config.saudi_cost_premium,
        "Outsource cost discount": config.outsource_cost_discount,
    }
    for key, note in COST_ASSUMPTIONS:
        v = value_lookup[key]
        ws.append([key, "" if v is None else v, note])
    for row in ws.iter_rows(min_row=2):
        row[2].font = _INSTRUCTION_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 100


# ──────────────────────────────────────────────────────────────────────────
# Helpers — parsing each sheet


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_example_value(value: str) -> bool:
    """Skeleton sheets include italic placeholder rows that start with '(example)'.
    The parser must ignore them so the consultant's empty BU doesn't end up with
    fake mapping data after a re-upload."""
    return value.startswith("(example)")


def _parse_profession_mapping(ws, config: BUConfigurationPayload, errors: list[str]):
    """Activity/profession mapping rows are user-defined free text per BU — we don't
    enforce the standardized value matches any canonical list. We just collect every
    row that has both a raw and a standardized value."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = _cell_text(row[0] if len(row) > 0 else "")
        standardized = _cell_text(row[1] if len(row) > 1 else "")
        if not raw or not standardized or _is_example_value(raw):
            continue
        config.profession_mapping[raw] = standardized


def _parse_activity_mapping(ws, config: BUConfigurationPayload, errors: list[str]):
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = _cell_text(row[0] if len(row) > 0 else "")
        standardized = _cell_text(row[1] if len(row) > 1 else "")
        if not raw or not standardized or _is_example_value(raw):
            continue
        config.activity_mapping[raw] = standardized


def _parse_job_families(ws, config: BUConfigurationPayload, errors: list[str]):
    """Job Families sheet: Activity | Profession | Job Family | Outsourceability.
    Populates BOTH ``job_family_mapping`` (pair → family) and ``outsourceability_overrides``
    (family → outsourceability) from the same rows."""
    family_outsourceability: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        activity = _cell_text(row[0] if len(row) > 0 else "")
        profession = _cell_text(row[1] if len(row) > 1 else "")
        family = _cell_text(row[2] if len(row) > 2 else "")
        outsourceability = _cell_text(row[3] if len(row) > 3 else "")
        if not family:
            continue
        if _is_example_value(activity) or _is_example_value(profession):
            continue
        if activity or profession:
            pair_key = f"{activity} - {profession}".strip(" -")
            if pair_key:
                config.job_family_mapping[pair_key] = family
        if outsourceability:
            if outsourceability not in VALID_OUTSOURCEABILITY:
                errors.append(
                    f"Job Families: family '{family}' outsourceability must be one of "
                    f"{sorted(VALID_OUTSOURCEABILITY)}, got '{outsourceability}'"
                )
                continue
            existing = family_outsourceability.get(family)
            if existing and existing != outsourceability:
                errors.append(
                    f"Job Families: family '{family}' has conflicting outsourceability values "
                    f"('{existing}' vs '{outsourceability}'). Use a single value per family."
                )
                continue
            family_outsourceability[family] = outsourceability
    config.outsourceability_overrides.update(family_outsourceability)


def _parse_outsourceability(ws, config: BUConfigurationPayload, errors: list[str]):
    """Legacy standalone Outsourceability sheet (PR-#2 era). Still honored when present;
    its rows merge into the same outsourceability_overrides dict."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        family = _cell_text(row[0] if len(row) > 0 else "")
        value = _cell_text(row[1] if len(row) > 1 else "")
        if not family or not value:
            continue
        if family not in OUTSOURCEABILITY_RULES:
            errors.append(f"Outsourceability: unknown job family '{family}' (typo?)")
            continue
        if value not in VALID_OUTSOURCEABILITY:
            errors.append(
                f"Outsourceability: '{family}' value must be one of {sorted(VALID_OUTSOURCEABILITY)}, got '{value}'"
            )
            continue
        config.outsourceability_overrides[family] = value


def _parse_ratios(ws, config: BUConfigurationPayload, errors: list[str]):
    for row in ws.iter_rows(min_row=2, values_only=True):
        family = _cell_text(row[0] if len(row) > 0 else "")
        value = _cell_text(row[1] if len(row) > 1 else "")
        if not family or not value:
            continue
        if family not in MAXIMUM_RATIO_RULES:
            errors.append(f"Ratios: unknown supervisor family '{family}' (typo?)")
            continue
        if not RATIO_PATTERN.match(value):
            errors.append(
                f"Ratios: '{family}' value must be of the form '1:N' (got '{value}')"
            )
            continue
        config.ratio_overrides[family] = value.replace(" ", "")


def _parse_drivers(ws, config: BUConfigurationPayload, errors: list[str]):
    valid_supervisors = set(DRIVER_SUPERVISORS)
    for row in ws.iter_rows(min_row=2, values_only=True):
        family = _cell_text(row[0] if len(row) > 0 else "")
        activity = _cell_text(row[1] if len(row) > 1 else "")
        profession = _cell_text(row[2] if len(row) > 2 else "")
        if not family:
            continue
        if family not in valid_supervisors:
            errors.append(f"Drivers: unknown supervisor family '{family}' (typo?)")
            continue
        if not activity and not profession:
            continue
        config.driver_overrides.setdefault(family, []).append({
            "activity": activity,
            "profession": profession,
        })


def _parse_cost_assumptions(ws, config: BUConfigurationPayload, errors: list[str]):
    """Parses both the new 'Cost Assumptions' sheet and the legacy 'Engine Knobs' sheet.
    Accepts both 'Saudi pay premium' and the legacy 'Saudi cost premium' label."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _cell_text(row[0] if len(row) > 0 else "")
        value_raw = row[1] if len(row) > 1 else None
        if not key:
            continue
        if isinstance(value_raw, str) and not value_raw.strip():
            value_raw = None
        if value_raw is None:
            continue
        try:
            value = float(value_raw)
        except (TypeError, ValueError):
            errors.append(f"Cost Assumptions: '{key}' value must be a number (got '{value_raw}')")
            continue
        key_normalized = key.lower().strip()
        if key_normalized in ("saudi pay premium", "saudi cost premium"):
            if value < 1.0 or value > 3.0:
                errors.append(
                    f"Cost Assumptions: 'Saudi pay premium' must be between 1.0 and 3.0 (got {value})"
                )
                continue
            config.saudi_cost_premium = value
        elif key_normalized == "outsource cost discount":
            if value < 0.0 or value > 1.0:
                errors.append(
                    f"Cost Assumptions: 'Outsource cost discount' must be between 0 and 1 (got {value})"
                )
                continue
            config.outsource_cost_discount = value
        else:
            errors.append(f"Cost Assumptions: unknown knob '{key}' (typo?)")
